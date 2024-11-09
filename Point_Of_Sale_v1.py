import tkinter as tk
from tkinter import ttk, messagebox, font
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
from calendar import month_name

# Ensure the Excel file is created
def create_product_excel(file_name="products.xlsx"):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["ID", "Name", "Price", "Quantity"])
        wb.save(file_name)

def create_sales_excel(file_name="sales.xlsx"):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Date", "Product ID", "Product Name", "Quantity", "Price", "Total",
                  "Discount", "GST", "Final Total", "Customer Name", "Customer Phone",
                  "Payment Type", "Amount Paid", "Amount Due"])
        wb.save(file_name)

def record_sale(sale_data, file_name="sales.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.append(sale_data)
    wb.save(file_name)

class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Login")
        self.root.geometry("300x150")
        self.root.configure(bg="#2C3E50")

        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.root, text="Username").grid(row=0, column=0, padx=10, pady=10)
        self.entry_username = ttk.Entry(self.root)
        self.entry_username.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.root, text="Password").grid(row=1, column=0, padx=10, pady=10)
        self.entry_password = ttk.Entry(self.root, show="*")
        self.entry_password.grid(row=1, column=1, padx=10, pady=10)

        ttk.Button(self.root, text="Login", command=self.login).grid(row=2, column=0, columnspan=2, pady=10)

    def login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()
        if username == "ss" and password == "4176":
            self.root.destroy()
            root = tk.Tk()
            app = POSApp(root)
            root.mainloop()
        else:
            messagebox.showerror("Error", "Invalid credentials")

def add_product(file_name, product_id, name, price, quantity):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    ws.append([product_id, name, price, quantity])
    wb.save(file_name)

def display_products(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def search_product(file_name, search_term):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if search_term.lower() in str(row[0]).lower() or search_term.lower() in str(row[1]).lower():
            results.append(row)
    return results

def update_product(file_name, product_id, new_name=None, new_price=None, new_quantity=None):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            if new_name:
                row[1].value = new_name
            if new_price:
                row[2].value = new_price
            if new_quantity:
                row[3].value = new_quantity
            wb.save(file_name)
            return True
    return False

def delete_product(file_name, product_id):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            ws.delete_rows(row[0].row, 1)
            wb.save(file_name)
            return True
    return False

def place_order_product(file_name, product_id, order_quantity):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[3].value += order_quantity
            wb.save(file_name)
            return True
    return False

def purchase_product(file_name, product_id, purchase_quantity):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            if row[3].value >= purchase_quantity:
                row[3].value -= purchase_quantity
                wb.save(file_name)
                return True, row[3].value
            else:
                return False, row[3].value
    return False, None

class POSApp:
    def __init__(self, root):
        self.root = root
        self.root.title("POS System")
        self.root.geometry("800x600")
        self.root.configure(bg="#2C3E50")
        # Add this after initializing root
        style = ttk.Style()
        style.configure('TButton',
                        padding=10,
                        font=('Helvetica', 10, 'bold'),
                        background='#2196F3'
                        )

        style.configure('TLabel',
                        font=('Helvetica', 10),
                        padding=5
                        )

        style.configure('Treeview',
                        font=('Helvetica', 9),
                        rowheight=25
                        )

        style.configure('TFrame',
                        background='#f5f5f5'
                        )

        self.file_name = "products.xlsx"
        self.sales_file_name = "sales.xlsx"
        self.purchase_items = []
        self.cart = []
        self.create_widgets()

    # In your POSApp class initialization
        self.root.resizable(True, True)
        self.root.minsize(800, 600)

        # Add this in your POSApp class
        for i in range(3):
            self.root.grid_columnconfigure(i, weight=1)
            self.root.grid_rowconfigure(i, weight=1)

        # For your treeviews
        self.tree_sales.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.cart_tree_view.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

    def create_widgets(self):
        self.custom_font = font.Font(family="Helvetica", size=12, weight="bold")
        style = ttk.Style()
        style.configure("TNotebook", background="#34495E", foreground="#ECF0F1", font=self.custom_font)
        style.configure("TFrame", background="#34495E")
        style.configure("TLabel", background="#34495E", foreground="#ECF0F1", font=self.custom_font)
        style.configure("TEntry", font=self.custom_font)
        style.configure("Treeview.Heading", background="#34495E", foreground="#34495E",font=self.custom_font)  # Dark background with black text
        style.configure("Treeview", background="#ECF0F1", foreground="#34495E",font=self.custom_font)  # Set Treeview background and text colors
        style.configure("TButton", background="#2980B9", foreground="#34495E",font=self.custom_font, padding=5)  # Button with black text
        style.map("TButton", background=[("active", "#3498DB")])  # Change button color on hover

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=20, expand=True, fill=tk.BOTH)

        self.tab_welcome = ttk.Frame(self.notebook)
        self.tab_add = ttk.Frame(self.notebook)
        self.tab_view = ttk.Frame(self.notebook)
        self.tab_purchase = ttk.Frame(self.notebook)
        self.tab_admin = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_welcome, text="Welcome")
        self.notebook.add(self.tab_add, text="Add Product")
        self.notebook.add(self.tab_view, text="View Products")
        self.notebook.add(self.tab_purchase, text="Purchase Product")
        self.notebook.add(self.tab_admin, text="Admin")

        self.create_welcome_tab()
        self.create_add_product_tab()
        self.create_view_products_tab()
        self.create_purchase_product_tab()
        self.create_admin_tab()

    def create_welcome_tab(self):
        welcome_label = ttk.Label(self.tab_welcome, text="Welcome to SS INTERPRISORS POS System", font=("Helvetica", 24))
        welcome_label.pack(pady=20)
        footer_label = ttk.Label(self.tab_welcome, text="Contact: +92 333 5130796\nEmail: chsajjadshahid@outlook.com\nAddress: I-10/MARKAZ ,ISLAMABAD", font=("Helvetica", 12), background="#34495E", foreground="#ECF0F1")
        footer_label.pack(side="bottom", pady=20)

    def create_add_product_tab(self):
        ttk.Label(self.tab_add, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_product_id = ttk.Entry(self.tab_add)
        self.entry_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Name").grid(row=1, column=0, padx=10, pady=10)
        self.entry_product_name = ttk.Entry(self.tab_add)
        self.entry_product_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Price").grid(row=2, column=0, padx=10, pady=10)
        self.entry_product_price = ttk.Entry(self.tab_add)
        self.entry_product_price.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Quantity").grid(row=3, column=0, padx=10, pady=10)
        self.entry_product_quantity = ttk.Entry(self.tab_add)
        self.entry_product_quantity.grid(row=3, column=1, padx=10, pady=10)

        ttk.Button(self.tab_add, text="Add Product", command=self.add_product).grid(row=4, column=0, columnspan=2, pady=10)

    def add_product(self):
        product_id = int(self.entry_product_id.get())
        name = self.entry_product_name.get()
        price = float(self.entry_product_price.get())
        quantity = int(self.entry_product_quantity.get())

        add_product(self.file_name, product_id, name, price, quantity)
        messagebox.showinfo("Success", "Product added successfully")
        self.clear_entries()


    def create_view_products_tab(self):
        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_view)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Treeview for displaying products
        self.tree_view = ttk.Treeview(tree_frame, columns=("ID", "Name", "Price", "Quantity"), show="headings")
        self.tree_view.heading("ID", text="ID")
        self.tree_view.heading("Name", text="Name")
        self.tree_view.heading("Price", text="Price")
        self.tree_view.heading("Quantity", text="Quantity")
        self.tree_view.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_view.yview)
        self.tree_view.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_view.xview)
        self.tree_view.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Configure the grid layout for tree_frame to allow resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Refresh button
        ttk.Button(self.tab_view, text="Refresh", command=self.refresh_view_products).pack(pady=10)



    def refresh_view_products(self):
        for row in self.tree_view.get_children():
            self.tree_view.delete(row)
        for row in display_products(self.file_name):
            self.tree_view.insert("", "end", values=row)

    def revert_quantity(self, file_name, product_id, quantity):
        wb = openpyxl.load_workbook(file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[3].value += quantity
                wb.save(file_name)
                return True
        return False

    def clear_cart(self):
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            self.revert_quantity(self.file_name, product_id, quantity)
        self.cart_tree = []
        self.cart_total = 0
        self.refresh_cart()
        self.cart_window.destroy()

    def create_purchase_product_tab(self):
        # Create main frames to organize content
        input_frame = ttk.Frame(self.tab_purchase)
        cart_frame = ttk.Frame(self.tab_purchase)

        input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        cart_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Product Information Section
        ttk.Label(input_frame, text="Product Information", font=("Helvetica", 12, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(0, 10))

        labels = ["Product ID", "Quantity", "Discount (%)", "GST (%)"]
        entries = ["entry_purchase_product_id", "entry_purchase_quantity", "entry_discount", "entry_gst"]

        for i, (label, entry) in enumerate(zip(labels, entries)):
            ttk.Label(input_frame, text=label).grid(row=i + 1, column=0, padx=5, pady=5, sticky="e")
            setattr(self, entry, ttk.Entry(input_frame))
            getattr(self, entry).grid(row=i + 1, column=1, padx=5, pady=5, sticky="ew")

        # Customer Information Section
        ttk.Label(input_frame, text="Customer Information", font=("Helvetica", 12, "bold")).grid(
            row=5, column=0, columnspan=2, pady=(20, 10))

        # Customer details
        customer_info = [
            ("Customer Name", "entry_customer_name"),
            ("Customer Phone", "entry_customer_phone"),
            ("Payment Type", "payment_type_combo"),
            ("Amount Paid", "entry_amount_paid")
        ]

        for i, (label, entry) in enumerate(customer_info):
            ttk.Label(input_frame, text=label).grid(row=i + 6, column=0, padx=5, pady=5, sticky="e")
            if label == "Payment Type":
                self.payment_type = tk.StringVar(value="Cash")
                self.payment_type_combo = ttk.Combobox(input_frame, textvariable=self.payment_type)
                self.payment_type_combo['values'] = ('Cash', 'Credit')
                self.payment_type_combo.grid(row=i + 6, column=1, padx=5, pady=5, sticky="ew")
            else:
                setattr(self, entry, ttk.Entry(input_frame))
                getattr(self, entry).grid(row=i + 6, column=1, padx=5, pady=5, sticky="ew")

        # Buttons
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=10, column=0, columnspan=2, pady=20)

        self.add_to_cart_button = ttk.Button(button_frame, text="Add to Cart", command=self.add_to_cart)
        self.add_to_cart_button.grid(row=0, column=0, padx=5)

        self.view_cart_button = ttk.Button(button_frame, text="View Cart", command=self.view_cart)
        self.view_cart_button.grid(row=0, column=1, padx=5)

        # Cart Section (right side)
        ttk.Label(cart_frame, text="Shopping Cart", font=("Helvetica", 12, "bold")).grid(
            row=0, column=0, pady=(0, 10))

        # Treeview for cart items
        self.cart_tree = []
        self.cart_total = 0

        tree_container = ttk.Frame(cart_frame)
        tree_container.grid(row=1, column=0, sticky="nsew")

        self.cart_tree_view = ttk.Treeview(
            tree_container,
            columns=("ID", "Name", "Price", "Quantity", "Base Total", "Discount", "GST", "Final Total"),
            show="headings"
        )

        # Configure column headings
        for col in self.cart_tree_view["columns"]:
            self.cart_tree_view.heading(col, text=col)
            self.cart_tree_view.column(col, width=100)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.cart_tree_view.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.cart_tree_view.xview)
        self.cart_tree_view.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Grid layout for treeview and scrollbars
        self.cart_tree_view.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Total section
        total_frame = ttk.Frame(cart_frame)
        total_frame.grid(row=2, column=0, pady=10, sticky="e")

        ttk.Label(total_frame, text="Total:").grid(row=0, column=0, padx=5)
        self.cart_total_label = ttk.Label(total_frame, text="0.00")
        self.cart_total_label.grid(row=0, column=1, padx=5)

        # Configure weights for proper resizing
        self.tab_purchase.grid_columnconfigure(1, weight=3)
        self.tab_purchase.grid_columnconfigure(0, weight=1)
        self.tab_purchase.grid_rowconfigure(0, weight=1)

        cart_frame.grid_columnconfigure(0, weight=1)
        cart_frame.grid_rowconfigure(1, weight=1)

        tree_container.grid_columnconfigure(0, weight=1)
        tree_container.grid_rowconfigure(0, weight=1)

    def add_to_cart(self):
        product_id = int(self.entry_purchase_product_id.get())
        quantity = int(self.entry_purchase_quantity.get())
        # Remove purchase_product call here since it's reducing stock immediately
        name = self.get_product_name(product_id)
        price = self.get_product_price(product_id)

        # Check if stock is available without updating it
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                current_stock = row[3].value
                if current_stock >= quantity:
                    # Calculate base price and totals with GST and discount
                    base_total = price * quantity
                    discount = float(self.entry_discount.get() if self.entry_discount.get() else 0)
                    gst = float(self.entry_gst.get() if self.entry_gst.get() else 0)

                    discount_amount = base_total * (discount / 100)
                    after_discount = base_total - discount_amount
                    gst_amount = after_discount * (gst / 100)
                    final_total = after_discount + gst_amount

                    # Store all values in cart_tree
                    self.cart_tree.append((product_id, name, price, quantity, base_total, discount, gst, final_total))
                    self.cart_total += final_total
                    self.refresh_cart()
                    messagebox.showinfo("Success", f"Added {quantity} items to cart. Available stock: {current_stock}")
                else:
                    messagebox.showerror("Error", f"Not enough stock. Available stock: {current_stock}")
                break
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def update_stock_quantity(self, product_id, quantity):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[3].value -= quantity
                wb.save(self.file_name)
                return True
        return False

    def view_cart(self):
        self.create_cart_window()

    def create_cart_window(self):
        self.cart_window = tk.Toplevel(self.root)
        self.cart_window.title("Shopping Cart")
        self.cart_window.geometry("800x700")
        self.cart_window.configure(bg='#f0f0f0')

        # For cart window
        self.cart_window.resizable(True, True)
        self.cart_window.minsize(800, 700)

        # Header
        header_frame = ttk.Frame(self.cart_window)
        header_frame.pack(fill='x', pady=10)
        ttk.Label(header_frame, text="Your Shopping Cart",
                  font=('Helvetica', 16, 'bold')).pack()

        # Cart content
        cart_frame = ttk.Frame(self.cart_window)
        cart_frame.pack(padx=20, pady=10, fill='both', expand=True)

        # Styled text widget
        cart_text = tk.Text(cart_frame, height=20, width=80,
                            font=('Courier', 10),
                            bg='white',
                            relief='ridge',
                            borderwidth=2)
        cart_text.pack(padx=10, pady=10)

        # Header formatting with colors
        cart_data = f"""
        {'ID':<10}{'Name':<20}{'Price':<10}{'Quantity':<10}{'Base Total':<12}{'Discount':<10}{'GST':<10}{'Final Total':<10}\n"""
        cart_text.insert(tk.END, cart_data)
        cart_text.tag_add("header", "2.0", "3.0")
        cart_text.tag_config("header", background="#4a90e2", foreground="white", font=('Courier', 10, 'bold'))

        # Cart items
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            item_data = f"{product_id:<10}{name:<20}{price:<10.2f}{quantity:<10}{base_total:<12.2f}{discount:<10}%{gst:<10}%{final_total:<10.2f}\n"
            cart_text.insert(tk.END, item_data)

        # Total amount with styling
        total_text = f"\nTotal Amount: ₹{self.cart_total:.2f}"
        cart_text.insert(tk.END, total_text)
        cart_text.tag_add("total", f"end-2c linestart", "end")
        cart_text.tag_config("total", font=('Helvetica', 12, 'bold'))

        # Buttons frame
        button_frame = ttk.Frame(self.cart_window)
        button_frame.pack(pady=20, padx=20)

        # Style configuration for buttons
        style = ttk.Style()
        style.configure('Action.TButton', padding=10, font=('Helvetica', 10))

        # Buttons with improved layout
        buttons = [
            ("Clear Cart", self.clear_cart, '#ff6b6b'),
            ("Checkout", self.checkout, '#51cf66'),
            ("Generate Bill", self.generate_bill, '#339af0'),
            ("Print Bill", self.print_bill, '#845ef7')
        ]

        for text, command, color in buttons:
            btn = ttk.Button(button_frame, text=text, command=command, style='Action.TButton')
            btn.pack(side='left', padx=10)

        # Make text widget read-only
        cart_text.configure(state='disabled')

    def refresh_cart(self):
        for row in self.cart_tree_view.get_children():
            self.cart_tree_view.delete(row)
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            self.cart_tree_view.insert("", "end", values=(product_id, name, price, quantity, base_total,
                                                          f"{discount}%", f"{gst}%", final_total))
        self.cart_total_label.config(text=f"{self.cart_total:.2f}")

    def checkout(self):
        if not self.cart_tree:
            messagebox.showerror("Error", "Cart is empty")
            return

        # Now update stock quantities only during checkout
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            self.update_stock_quantity(product_id, quantity)

        self.generate_bill()
        self.record_sales()
        self.cart_tree = []
        self.cart_total = 0
        self.refresh_cart()
        self.cart_window.destroy()

    def record_sales(self):
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            customer_name = self.entry_customer_name.get()
            customer_phone = self.entry_customer_phone.get()
            payment_type = self.payment_type.get()
            amount_paid = float(self.entry_amount_paid.get() or 0)
            amount_due = final_total - amount_paid if payment_type == "Credit" else 0

            sale_data = [
                date, product_id, name, quantity, price, base_total,
                discount, gst, final_total,
                customer_name, customer_phone,
                payment_type, amount_paid, amount_due
            ]
            record_sale(sale_data, self.sales_file_name)

    def generate_bill(self):
        self.bill_window = tk.Toplevel(self.root)
        self.bill_window.title("Bill Receipt")
        self.bill_window.configure(bg='#f0f0f0')

        # Set window size and center it
        window_width = 600
        window_height = 800
        screen_width = self.bill_window.winfo_screenwidth()
        screen_height = self.bill_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.bill_window.geometry(f'{window_width}x{window_height}+{x}+{y}')

        # Create main frame
        main_frame = ttk.Frame(self.bill_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Configure Text widget with better styling
        self.bill_text = tk.Text(
            main_frame,
            height=30,
            width=80,
            font=('Courier', 10),
            bg='white',
            relief='solid'
        )
        self.bill_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Create the bill content with better formatting
        bill_data = f"""
    ╔══════════════════════════════════════════════════════════════╗
                            SS INTERPRISORS
                      Address: I-10/MARKAZ, ISLAMABAD
                        Contact: +92 333 5130796
    ╚══════════════════════════════════════════════════════════════╝

    Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    Customer Name: {self.entry_customer_name.get()}
    Customer Phone: {self.entry_customer_phone.get()}
    Payment Type: {self.payment_type.get()}

    {'═' * 60}
       PRODUCT DETAILS
    {'═' * 60}
    """

        # Add items with better alignment
        subtotal = 0
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            bill_data += f"│ {name:<20} │ Price: {price:>8} │ Qty: {quantity:>3} │ Total: {final_total:>8.2f} │\n"
            subtotal += base_total

        # Calculate totals
        discount = float(self.entry_discount.get() or 0)
        gst = float(self.entry_gst.get() or 0)
        discount_amount = subtotal * (discount / 100)
        amount_after_discount = subtotal - discount_amount
        gst_amount = amount_after_discount * (gst / 100)
        final_total = amount_after_discount + gst_amount
        amount_paid = float(self.entry_amount_paid.get() or 0)
        amount_due = final_total - amount_paid

        # Add summary with better formatting
        bill_data += f"""
    {'═' * 60}
                            BILL SUMMARY
    {'═' * 60}
        Subtotal:                               {subtotal:>10.2f}
        Discount ({discount}%):                         {discount_amount:>10.2f}
        GST ({gst}%):                                {gst_amount:>10.2f}
        ────────────────────────────────────────────────────
        Final Total:                            {final_total:>10.2f}
        Amount Paid:                            {amount_paid:>10.2f}
        Amount Due:                             {amount_due:>10.2f}
    {'═' * 60}

                Thank you for your business!
                Please visit again.
    """

        self.bill_text.insert(tk.END, bill_data)

        # Create a styled print button
        style = ttk.Style()
        style.configure('Print.TButton', font=('Arial', 10, 'bold'))
        print_btn = ttk.Button(
            main_frame,
            text="Print Bill",
            style='Print.TButton',
            command=self.print_bill
        )
        print_btn.pack(pady=10)


    def print_bill(self):
        bill_text = self.bill_text.get(1.0, tk.END)
        print(bill_text)
        # Optionally, you could implement printing to a physical printer using external libraries or OS-specific commands.

    def get_product_name(self, product_id):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                return row[1]
        return ""

    def get_product_price(self, product_id):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                return row[2]
        return 0
#################################new purchase###############
    def create_admin_tab(self):
        self.admin_notebook = ttk.Notebook(self.tab_admin)
        self.admin_notebook.pack(pady=20, expand=True, fill=tk.BOTH)

        self.tab_search = ttk.Frame(self.admin_notebook)
        self.tab_update = ttk.Frame(self.admin_notebook)
        self.tab_delete = ttk.Frame(self.admin_notebook)
        self.tab_order = ttk.Frame(self.admin_notebook)
        self.tab_sales_report = ttk.Frame(self.admin_notebook)  # New Sales Report tab

        self.admin_notebook.add(self.tab_search, text="Search Product")
        self.admin_notebook.add(self.tab_update, text="Update Product")
        self.admin_notebook.add(self.tab_delete, text="Delete Product")
        self.admin_notebook.add(self.tab_order, text="Place Order")
        self.admin_notebook.add(self.tab_sales_report, text="Sales Report")  # Adding Sales Report tab

        self.create_search_product_tab()
        self.create_update_product_tab()
        self.create_delete_product_tab()
        self.create_order_product_tab()
        self.create_sales_report_tab()  # Call to create the sales report tab

    def create_sales_report_tab(self):
        # Month selection
        ttk.Label(self.tab_sales_report, text="Select Month", font=("Arial", 10, "bold"), foreground="white",
                  background="#2d3e50").grid(row=0, column=0, padx=10, pady=10, sticky="W")

        self.month_var = tk.StringVar()
        self.month_dropdown = ttk.Combobox(self.tab_sales_report, textvariable=self.month_var, width=15)
        self.month_dropdown['values'] = [month_name[i] for i in range(1, 13)]
        self.month_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky="W")

        # Year selection
        ttk.Label(self.tab_sales_report, text="Select Year", font=("Arial", 10, "bold"), foreground="white",
                  background="#2d3e50").grid(row=0, column=2, padx=10, pady=10, sticky="W")

        self.year_var = tk.StringVar()
        self.year_dropdown = ttk.Combobox(self.tab_sales_report, textvariable=self.year_var, width=15)
        self.year_dropdown['values'] = [str(year) for year in range(2000, datetime.now().year + 1)]
        self.year_dropdown.grid(row=0, column=3, padx=10, pady=10, sticky="W")

        # Filter Button
        filter_button = ttk.Button(self.tab_sales_report, text="Filter", command=self.filter_sales_report)
        filter_button.grid(row=0, column=4, padx=10, pady=10, sticky="W")

        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_sales_report)
        tree_frame.grid(row=1, column=0, columnspan=5, padx=10, pady=20, sticky="NSEW")

        # Treeview for displaying sales
        self.tree_sales = ttk.Treeview(tree_frame, columns=(
            "Date", "Product ID", "Product Name", "Quantity", "Price", "Total",
            "Discount", "GST", "Final Total", "Customer Name", "Customer Phone",
            "Payment Type", "Amount Paid", "Amount Due"), show="headings")
        # Set column headings and widths
        for col in self.tree_sales["columns"]:
            self.tree_sales.heading(col, text=col)
            if col in ["Discount", "GST", "Price", "Total", "Final Total", "Amount Paid", "Amount Due"]:
                self.tree_sales.column(col, width=100, anchor="e")  # Right-align numeric columns
            elif col in ["Date", "Product ID"]:
                self.tree_sales.column(col, width=120)
            else:
                self.tree_sales.column(col, width=150)

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_sales.yview)
        self.tree_sales.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_sales.xview)
        self.tree_sales.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Position the Treeview in the frame
        self.tree_sales.grid(row=0, column=0, sticky="nsew")

        # Configure tree_frame to be resizable
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure grid expansion for resizing
        self.tab_sales_report.columnconfigure(4, weight=1)
        self.tab_sales_report.rowconfigure(1, weight=1)

        # Load all sales initially
        self.load_sales_data()

    def load_sales_data(self):
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)

        wb = openpyxl.load_workbook(self.sales_file_name)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Convert numeric values to float and format them
            formatted_row = list(row)
            for i, value in enumerate(row):
                # Format numeric columns (Price, Total, Discount, GST, etc.)
                if i in [4, 5, 6, 7, 8, 12, 13]:  # Indices for numeric columns
                    try:
                        formatted_row[i] = float(value) if value is not None else 0.0
                    except (ValueError, TypeError):
                        formatted_row[i] = 0.0

            self.tree_sales.insert("", "end", values=formatted_row)
    def filter_sales_report(self):
        """Filters sales based on selected month and year"""
        selected_month = self.month_var.get()
        selected_year = self.year_var.get()

        if not selected_month or not selected_year:
            messagebox.showerror("Error", "Please select both month and year.")
            return

        month_num = list(month_name).index(selected_month)  # Convert month name to month number
        filtered_sales = []

        wb = openpyxl.load_workbook(self.sales_file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            sale_date = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")

            if sale_date.month == month_num and sale_date.year == int(selected_year):
                filtered_sales.append(row)

        # Clear the tree view and insert filtered data
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)
        for row in filtered_sales:
            self.tree_sales.insert("", "end", values=row)
    def create_search_product_tab(self):
        ttk.Label(self.tab_search, text="Search Term").grid(row=0, column=0, padx=10, pady=10)
        self.entry_search_term = ttk.Entry(self.tab_search)
        self.entry_search_term.grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(self.tab_search, text="Search", command=self.search_product).grid(row=1, column=0, columnspan=2,
                                                                                     pady=10)

        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_search)
        tree_frame.grid(row=2, column=0, columnspan=2, pady=10, padx=10, sticky="nsew")

        # Treeview for displaying search results
        self.tree_search = ttk.Treeview(tree_frame, columns=("ID", "Name", "Price", "Quantity"), show="headings")
        self.tree_search.heading("ID", text="ID")
        self.tree_search.heading("Name", text="Name")
        self.tree_search.heading("Price", text="Price")
        self.tree_search.heading("Quantity", text="Quantity")
        self.tree_search.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_search.yview)
        self.tree_search.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_search.xview)
        self.tree_search.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Configure the grid layout for tree_frame to allow resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure the grid expansion for the main tab to allow resizing
        self.tab_search.columnconfigure(1, weight=1)
        self.tab_search.rowconfigure(2, weight=1)

    def search_product(self):
        search_term = self.entry_search_term.get()
        for row in self.tree_search.get_children():
            self.tree_search.delete(row)
        for row in search_product(self.file_name, search_term):
            self.tree_search.insert("", "end", values=row)

    def create_update_product_tab(self):
        ttk.Label(self.tab_update, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_update_product_id = ttk.Entry(self.tab_update)
        self.entry_update_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_update, text="New Name").grid(row=1, column=0, padx=10, pady=10)
        self.entry_update_name = ttk.Entry(self.tab_update)
        self.entry_update_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.tab_update, text="New Price").grid(row=2, column=0, padx=10, pady=10)
        self.entry_update_price = ttk.Entry(self.tab_update)
        self.entry_update_price.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.tab_update, text="New Quantity").grid(row=3, column=0, padx=10, pady=10)
        self.entry_update_quantity = ttk.Entry(self.tab_update)
        self.entry_update_quantity.grid(row=3, column=1, padx=10, pady=10)

        ttk.Button(self.tab_update, text="Update", command=self.update_product).grid(row=4, column=0, columnspan=2, pady=10)

    def update_product(self):
        product_id = int(self.entry_update_product_id.get())
        new_name = self.entry_update_name.get() or None
        new_price = self.entry_update_price.get()
        new_price = float(new_price) if new_price else None
        new_quantity = self.entry_update_quantity.get()
        new_quantity = int(new_quantity) if new_quantity else None

        if update_product(self.file_name, product_id, new_name, new_price, new_quantity):
            messagebox.showinfo("Success", "Product updated successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def create_delete_product_tab(self):
        ttk.Label(self.tab_delete, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_delete_product_id = ttk.Entry(self.tab_delete)
        self.entry_delete_product_id.grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(self.tab_delete, text="Delete", command=self.delete_product).grid(row=1, column=0, columnspan=2, pady=10)

    def delete_product(self):
        product_id = int(self.entry_delete_product_id.get())
        if delete_product(self.file_name, product_id):
            messagebox.showinfo("Success", "Product deleted successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def create_order_product_tab(self):
        ttk.Label(self.tab_order, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_order_product_id = ttk.Entry(self.tab_order)
        self.entry_order_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_order, text="Order Quantity").grid(row=1, column=0, padx=10, pady=10)
        self.entry_order_quantity = ttk.Entry(self.tab_order)
        self.entry_order_quantity.grid(row=1, column=1, padx=10, pady=10)

        ttk.Button(self.tab_order, text="Place Order", command=self.place_order_product).grid(row=2, column=0, columnspan=2, pady=10)

    def place_order_product(self):
        product_id = int(self.entry_order_product_id.get())
        order_quantity = int(self.entry_order_quantity.get())
        if place_order_product(self.file_name, product_id, order_quantity):
            messagebox.showinfo("Success", "Order placed successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def create_bill_window(self):
        self.bill_window = tk.Toplevel(self.root)
        self.bill_window.title("Bill Receipt")

        # For bill window
        self.bill_window.resizable(True, True)
        self.bill_window.minsize(600, 400)
        self.bill_text = tk.Text(self.bill_window, height=20, width=80)
        self.bill_text.pack(padx=20, pady=20)

        ttk.Button(self.bill_window, text="Print", command=self.print_bill).pack(pady=10)
        ttk.Button(self.bill_window, text="Generate Bill", command=self.generate_bill).pack(pady=10)


    def clear_entries(self):
        self.entry_product_id.delete(0, tk.END)
        self.entry_product_name.delete(0, tk.END)
        self.entry_product_price.delete(0, tk.END)
        self.entry_product_quantity.delete(0, tk.END)
        self.entry_purchase_product_id.delete(0, tk.END)
        self.entry_purchase_quantity.delete(0, tk.END)
        self.entry_discount.delete(0, tk.END)
        self.entry_gst.delete(0, tk.END)
        self.entry_search_term.delete(0, tk.END)
        self.entry_update_product_id.delete(0, tk.END)
        self.entry_update_name.delete(0, tk.END)
        self.entry_update_price.delete(0, tk.END)
        self.entry_update_quantity.delete(0, tk.END)
        self.entry_delete_product_id.delete(0, tk.END)
        self.entry_order_product_id.delete(0, tk.END)
        self.entry_order_quantity.delete(0, tk.END)

if __name__ == "__main__":
    create_product_excel()
    create_sales_excel()

    root = tk.Tk()
    app = LoginWindow(root)
    root.mainloop()

