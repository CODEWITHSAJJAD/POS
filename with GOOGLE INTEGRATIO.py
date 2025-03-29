from tkinter import ttk, font
from tkinter.ttk import Notebook
import openpyxl
from datetime import datetime
import sys
from calendar import month_name
####################
#backup iters
######################
import os
import tkinter as tk
from tkinter import messagebox
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Google API Scopes - Defined as a tuple to prevent modification
SCOPES = (
    'https://www.googleapis.com/auth/drive.file',
    'https://www.googleapis.com/auth/spreadsheets'
)

# OAuth Client Configuration
CLIENT_CONFIG = {
    "installed": {
        "client_id": "CLIENT id",
        "project_id": "project id",
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        "client_secret": "the client seret",
        "redirect_uris": ["http://localhost:8080"]
    }
}


########################################

def hide(self, tab_id):
    self.tab(tab_id, state='hidden')

def show(self, tab_id):
    self.tab(tab_id, state='normal')

Notebook.hide = hide
Notebook.show = show

# # Ensure the Excel file is created
WORKBOOK_PATH = "POS.XLSX"

def initialize_workbook():
    """
    Initialize the workbook by checking if it exists, creating it if needed,
    and ensuring all required sheets are present.
    """
    try:
        # Try to load existing workbook
        if os.path.exists(WORKBOOK_PATH):
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        else:
            # Create new workbook if it doesn't exist
            workbook = openpyxl.Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)

        # Required sheets and their headers
        required_sheets = {
            "products":["ID", "Name", "Price", "Quantity"],
            "sales":["Date", "Product ID", "Product Name", "Quantity", "Price", "Total",
                  "Discount", "GST", "Final Total", "Customer Name", "Customer Phone",
                  "Payment Type", "Amount Paid", "Amount Due"],
            "users":["Full Name", "Contact", "Username", "User Type", "Password"]

        }

        # Create or verify each required sheet
        for sheet_name, headers in required_sheets.items():
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                # Add headers
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)

                # If it's the Users sheet, add default admin account
                if sheet_name == "Users":
                    sheet.append(["admin", "admin123", "Admin"])

        # Save the workbook
        workbook.save(WORKBOOK_PATH)
        return workbook

    except Exception as e:
        messagebox.showerror("Error", f"Failed to initialize workbook: {e}")
        return None
workbook = initialize_workbook()
if workbook is None:
    sys.exit(1)  # Exit if workbook initialization fails
def record_sale(sale_data):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["sales"]
    ws.append(sale_data)
    wb.save(WORKBOOK_PATH)


class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Login")
        self.root.geometry("300x150")
        self.root.configure(bg="#2C3E50")

        # Create users.xlsx if it doesn't exist


        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.root, text="Username").grid(row=0, column=0, padx=10, pady=10)
        self.entry_username = ttk.Entry(self.root)
        self.entry_username.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.root, text="Password").grid(row=1, column=0, padx=10, pady=10)
        self.entry_password = ttk.Entry(self.root, show="*")
        self.entry_password.grid(row=1, column=1, padx=10, pady=10)

        ttk.Button(self.root, text="Login", command=self.login).grid(row=2, column=0, columnspan=2, pady=10)

    def login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()

        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["users"]

            user_found = False
            user_type = None
            full_name = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == username and row[4] == password:
                    user_found = True
                    user_type = row[3]
                    full_name = row[0]
                    break

            if user_found:
                self.root.destroy()
                root = tk.Tk()
                app = POSApp(root,username, user_type, full_name)

                # Configure access based on user type
                if user_type.lower() == "cashier":
                    # Hide all tabs except Welcome, Purchase, and View Products
                    app.notebook.hide(app.notebook.index(app.tab_add))
                    app.notebook.hide(app.notebook.index(app.tab_admin))

                elif user_type.lower() == "manager":

                    # For manager, hide only the user management section in admin tab
                    app.user_management_notebook.hide(app.user_management_notebook.index(app.tab_add_user))
                    app.user_management_notebook.hide(app.user_management_notebook.index(app.tab_delete_user))
                    app.user_management_notebook.hide(app.user_management_notebook.index(app.tab_update_user))
                    app.user_management_notebook.hide(app.user_management_notebook.index(app.tab_show_all_users))
                # Admin gets full access by default

                root.mainloop()
            else:
                messagebox.showerror("Error", "Invalid credentials")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


def add_product(product_id, name, price, quantity):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    ws.append([product_id, name, price, quantity])
    wb.save(WORKBOOK_PATH)

def display_products():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def search_product( search_term):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if search_term.lower() in str(row[0]).lower() or search_term.lower() in str(row[1]).lower():
            results.append(row)
    return results

def update_product( product_id, new_name=None, new_price=None, new_quantity=None):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            if new_name:
                row[1].value = new_name
            if new_price:
                row[2].value = new_price
            if new_quantity:
                row[3].value = new_quantity
            wb.save(WORKBOOK_PATH)
            return True
    return False

def delete_product( product_id):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            ws.delete_rows(row[0].row, 1)
            wb.save(WORKBOOK_PATH)
            return True
    return False

def place_order_product( product_id, order_quantity):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[3].value += order_quantity
            wb.save(WORKBOOK_PATH)
            return True
    return False

def purchase_product( product_id, purchase_quantity):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            if row[3].value >= purchase_quantity:
                row[3].value -= purchase_quantity
                wb.save(WORKBOOK_PATH)
                return True, row[3].value
            else:
                return False, row[3].value
    return False, None

class POSApp:
    def __init__(self, root, username, user_type, full_name):  # Modified constructor
        self.root = root
        self.root.title("POS System")
        self.root.geometry("800x600")
        self.root.configure(bg="#2C3E50")

        # Create top frame for user details and logout
        self.create_top_frame(username, user_type, full_name)
        # Add this after initializing root
        style = ttk.Style()
        style.configure('TButton',
                        padding=10,
                        font=('Helvetica', 10, 'bold'),
                        background='#2196F3'
                        )

        style.configure('TLabel',
                        font=('Helvetica', 10),
                        padding=5
                        )

        style.configure('Treeview',
                        font=('Helvetica', 9),
                        rowheight=25
                        )

        style.configure('TFrame',
                        background='#f5f5f5'
                        )
        self.purchase_items = []
        self.cart = []
        self.create_widgets()

    # In your POSApp class initialization
        self.root.resizable(True, True)
        self.root.minsize(800, 600)

        # Add this in your POSApp class
        for i in range(3):
            self.root.grid_columnconfigure(i, weight=1)
            self.root.grid_rowconfigure(i, weight=1)

        # For your treeviews
        self.tree_sales.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.cart_tree_view.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

    def create_top_frame(self, username, user_type, full_name):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill='x', padx=5, pady=5)

        # User details
        user_info = ttk.Label(
            top_frame,
            text=f"User: {full_name} | Type: {user_type}",
            font=('Helvetica', 10, 'bold')
        )
        user_info.pack(side='left', padx=10)

        # Logout button
        logout_btn = ttk.Button(
            top_frame,
            text="Logout",
            command=self.logout
        )
        logout_btn.pack(side='right', padx=10)

        self.GOOGLE_button = ttk.Button(top_frame, text="Google Login", command=self.G_login)
        self.GOOGLE_button.pack(side='right', padx=10)
        self.GOOGLE_button.config(state='normal')
        self.status_label = tk.Label(top_frame, text="Status: Not authenticated", bg='#34495E',fg="red")
        self.status_label.pack(side='right',pady=10)
        self.backup_button = ttk.Button(top_frame, text="Backup", command=self.backup_to_google)
        self.backup_button.pack(side='right', padx=10)
        self.backup_button.config(state='disabled')  # Initially disabled

    ############################################
    def check_existing_token(self):
        """Check for existing valid token."""
        if os.path.exists('token.json'):
            try:
                self.creds = Credentials.from_authorized_user_file('token.json', SCOPES)
                if self.creds and self.creds.valid:
                    self.enable_buttons()
                    self.status_label.config(text="Status: Authenticated",bg='#f5f5f5', fg="green")
                else:
                    os.remove('token.json')
            except Exception as e:
                messagebox.showwarning("Warning", f"Error loading credentials: {str(e)}")
                os.remove('token.json')

    def G_login(self):
        """Complete OAuth authentication flow with scope consistency."""
        try:
            # Clear existing credentials
            if os.path.exists('token.json'):
                os.remove('token.json')

            flow = InstalledAppFlow.from_client_config(
                CLIENT_CONFIG,
                scopes=SCOPES,
                redirect_uri='http://localhost:8080'
            )

            # Run with all requested scopes
            self.creds = flow.run_local_server(
                port=8080,
                prompt='consent',
                open_browser=True,
                authorization_prompt_message="",
                success_message='Authentication complete!',
                include_granted_scopes='true'
            )

            granted_scopes = self.creds.scopes
            for scope in SCOPES:
                if scope not in granted_scopes:
                    raise ValueError(f"Required scope {scope} was not granted")

            # Save credentials
            with open('token.json', 'w') as token:
                token.write(self.creds.to_json())

            self.GOOGLE_button.config(state='disabled')
            self.backup_button.config(state='normal')  # Enable
            self.status_label.config(text="Status: Authenticated", fg="green")
            messagebox.showinfo("Success", "Logged in successfully with all required permissions!")

        except Exception as e:
            messagebox.showerror("Login Failed", f"Error during login: {str(e)}")

            self.status_label.config(text="Status: Authentication failed", fg="red")
            if os.path.exists('token.json'):
                os.remove('token.json')

    def backup_to_google(self):
        """Backup all sheets to Google Sheets in a single workbook named 'POS'."""
        try:
            # Build both Sheets and Drive services
            sheets_service = build('sheets', 'v4', credentials=self.creds)
            drive_service = build('drive', 'v3', credentials=self.creds)  # Add Drive service

            # Define backup parameters
            spreadsheet_name = "POS"  # Name of the Google Sheets workbook
            sheets_to_backup = ['products', 'sales', 'users']  # Sheets to backup
            local_file = 'POS.xlsx'  # Local workbook file name

            # Get or create the spreadsheet
            spreadsheet_id = self.get_or_create_spreadsheet(drive_service, sheets_service, spreadsheet_name)

            if not spreadsheet_id:
                raise Exception("Failed to create or access Google Sheets workbook")

            # Backup each sheet
            for sheet_title in sheets_to_backup:
                self.backup_sheet(sheets_service, spreadsheet_id, local_file, sheet_title)

            messagebox.showinfo("Backup", "Backup completed successfully!")
        except Exception as e:
            messagebox.showerror("Backup Failed", f"Error during backup: {str(e)}")

    def get_or_create_spreadsheet(self, drive_service, sheets_service, title):
        """Get or create a Google Sheets spreadsheet with the required sheets."""
        try:
            # Search for existing spreadsheet using Drive API
            results = drive_service.files().list(
                q=f"name='{title}' and mimeType='application/vnd.google-apps.spreadsheet'",
                spaces='drive',
                fields='files(id, name)',
                pageSize=1
            ).execute()

            files = results.get('files', [])

            if files:
                # Spreadsheet exists, return its ID
                return files[0]['id']
            else:
                # Create new spreadsheet with all required sheets
                spreadsheet_body = {
                    'properties': {
                        'title': title
                    },
                    'sheets': [
                        {'properties': {'title': 'products'}},
                        {'properties': {'title': 'sales'}},
                        {'properties': {'title': 'users'}}
                    ]
                }

                spreadsheet = sheets_service.spreadsheets().create(
                    body=spreadsheet_body,
                    fields='spreadsheetId'
                ).execute()
                return spreadsheet.get('spreadsheetId')

        except HttpError as error:
            print(f"An error occurred: {error}")
            return None

    def backup_sheet(self, service, spreadsheet_id, file_name, sheet_title):
        """Backup a local Excel sheet to Google Sheets, handling datetime objects."""
        try:
            # Load local workbook and sheet
            wb = openpyxl.load_workbook(file_name)

            # Check if sheet exists in local file
            if sheet_title not in wb.sheetnames:
                print(f"Warning: Sheet '{sheet_title}' not found in local workbook")
                return

            ws = wb[sheet_title]

            # Convert cells to values, handling datetime objects
            values = []
            for row in ws.iter_rows():
                row_values = []
                for cell in row:
                    if isinstance(cell.value, datetime):  # Changed from datetime.datetime
                        # Format datetime as string (you can customize the format)
                        row_values.append(cell.value.strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        row_values.append(cell.value)
                values.append(row_values)

            # Clear existing data in the sheet
            service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range=sheet_title,
                body={}
            ).execute()

            # Update the sheet with new data
            body = {
                'values': values,
                'majorDimension': 'ROWS'
            }

            result = service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"{sheet_title}!A1",
                valueInputOption='USER_ENTERED',
                body=body
            ).execute()

            print(f"{result.get('updatedCells')} cells updated in {sheet_title}")

        except HttpError as error:
            print(f"An error occurred with sheet {sheet_title}: {error}")
            if error.resp.status == 404:
                self.create_sheet(service, spreadsheet_id, sheet_title)
                # Retry the backup
                self.backup_sheet(service, spreadsheet_id, file_name, sheet_title)
            else:
                raise

    def create_sheet(self, service, spreadsheet_id, sheet_title):
        """Create a new sheet in the Google Sheets workbook."""
        try:
            sheet_body = {
                'requests': [{
                    'addSheet': {
                        'properties': {
                            'title': sheet_title
                        }
                    }
                }]
            }
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=sheet_body
            ).execute()
            print(f"Created new sheet: {sheet_title}")
        except HttpError as error:
            print(f"Failed to create sheet {sheet_title}: {error}")
            raise
    def G_logout(self):
        """Log out by deleting the token file."""
        try:
            if os.path.exists('token.json'):
                os.remove('token.json')
            self.creds = None

            self.status_label.config(text="Status: Not authenticated", fg="red")
            messagebox.showinfo("Success", "Logged out successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to logout: {str(e)}")

    #######################################################################
    def logout(self):
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self.root.destroy()
            root = tk.Tk()
            app = LoginWindow(root)
            root.mainloop()
    def create_widgets(self):
        self.custom_font = font.Font(family="Helvetica", size=12, weight="bold")
        style = ttk.Style()
        style.configure("TNotebook", background="#34495E", foreground="#ECF0F1", font=self.custom_font)
        style.configure("TFrame", background="#34495E")
        style.configure("TLabel", background="#34495E", foreground="#ECF0F1", font=self.custom_font)
        style.configure("TEntry", font=self.custom_font)
        style.configure("Treeview.Heading", background="#34495E", foreground="#34495E",font=self.custom_font)  # Dark background with black text
        style.configure("Treeview", background="#ECF0F1", foreground="#34495E",font=self.custom_font)  # Set Treeview background and text colors
        style.configure("TButton", background="#2980B9", foreground="#34495E",font=self.custom_font, padding=5)  # Button with black text
        style.map("TButton", background=[("active", "#3498DB")])  # Change button color on hover

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=20, expand=True, fill=tk.BOTH)

        self.tab_welcome = ttk.Frame(self.notebook)
        self.tab_add = ttk.Frame(self.notebook)
        self.tab_view = ttk.Frame(self.notebook)
        self.tab_purchase = ttk.Frame(self.notebook)
        self.tab_admin = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_welcome, text="🏠 Welcome")
        self.notebook.add(self.tab_add, text="➕ Add Product")
        self.notebook.add(self.tab_view, text="📋 View Products")
        self.notebook.add(self.tab_purchase, text="🛒 Purchase")
        self.notebook.add(self.tab_admin, text="⚙️ Admin")

        self.create_welcome_tab()
        self.create_add_product_tab()
        self.create_view_products_tab()
        self.create_purchase_product_tab()
        self.create_admin_tab()

################################################################################
################welcome tab#####################################################
    def create_welcome_tab(self):
        # Create a main frame with a gradient-like effect
        main_frame = ttk.Frame(self.tab_welcome)
        main_frame.pack(expand=True, fill='both')

        # Style configuration
        style = ttk.Style()
        style.configure("Welcome.TLabel",
                        font=("Helvetica", 32, "bold"),
                        foreground="#2E86C1")

        style.configure("Footer.TLabel",
                        font=("Helvetica", 12),
                        foreground="#ECF0F1",
                        background="#34495E")

        # Company logo/image (assuming you have a logo.png in your assets)
        try:
            logo = tk.PhotoImage(file="assets/logo.png")
            logo_label = ttk.Label(main_frame, image=logo)
            logo_label.image = logo  # Keep a reference
            logo_label.pack(pady=20)
        except:
            pass  # Skip if no logo file exists

        # Welcome message with better styling
        welcome_frame = ttk.Frame(main_frame)
        welcome_frame.pack(pady=30)

        welcome_label = ttk.Label(welcome_frame,
                                  text="Welcome to",
                                  style="Welcome.TLabel")
        welcome_label.pack()

        company_label = ttk.Label(welcome_frame,
                                  text="SS INTERPRISORS",
                                  font=("Arial", 40, "bold"),
                                  foreground="#E74C3C")
        company_label.pack()

        pos_label = ttk.Label(welcome_frame,
                              text="Point of Sale System",
                              font=("Helvetica", 24),
                              foreground="#2C3E50")
        pos_label.pack()

        # Footer with contact information in a modern card-like design
        footer_frame = ttk.Frame(main_frame, style="Footer.TFrame")
        footer_frame.pack(side="bottom", fill="x", pady=20)

        contact_info = {
            "📞 Contact": "+92 333 5130796",
            "📧 Email": "chsajjadshahid@outlook.com",
            "📍 Address": "I-10/MARKAZ, ISLAMABAD"
        }

        for title, value in contact_info.items():
            info_frame = ttk.Frame(footer_frame)
            info_frame.pack(pady=5)

            ttk.Label(info_frame,
                      text=f"{title}: {value}",
                      style="Footer.TLabel").pack()

        # Add a simple animation effect (blinking cursor)
        def blink_cursor():
            current_color = company_label.cget("foreground")
            new_color = "#E74C3C" if current_color == "#2C3E50" else "#2C3E50"
            company_label.configure(foreground=new_color)
            self.tab_welcome.after(1000, blink_cursor)

        blink_cursor()

################################################################################
################add product tab and its functionality##########################
    def create_add_product_tab(self):
        ttk.Label(self.tab_add, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_product_id = ttk.Entry(self.tab_add)
        self.entry_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Name").grid(row=1, column=0, padx=10, pady=10)
        self.entry_product_name = ttk.Entry(self.tab_add)
        self.entry_product_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Price").grid(row=2, column=0, padx=10, pady=10)
        self.entry_product_price = ttk.Entry(self.tab_add)
        self.entry_product_price.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Quantity").grid(row=3, column=0, padx=10, pady=10)
        self.entry_product_quantity = ttk.Entry(self.tab_add)
        self.entry_product_quantity.grid(row=3, column=1, padx=10, pady=10)

        ttk.Button(self.tab_add, text="Add Product", command=self.add_product).grid(row=4, column=0, columnspan=2, pady=10)

    def add_product(self):
        product_id = int(self.entry_product_id.get())
        name = self.entry_product_name.get()
        price = float(self.entry_product_price.get())
        quantity = int(self.entry_product_quantity.get())

        # Read the existing products from the Excel file
        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["products"]

            # Check if the product ID already exists
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == product_id:
                    messagebox.showerror("Error",
                                         "Unable to add product. Product ID already exists. Please change the product ID.")
                    return

            # If the product ID is unique, proceed to add the product
            ws.append([product_id, name, price, quantity])
            wb.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Product added successfully")
            self.clear_entries()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    ################################################################################
##################view product tab and its functionality########################
    def create_view_products_tab(self):
        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_view)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Treeview for displaying products
        self.tree_view = ttk.Treeview(tree_frame, columns=("ID", "Name", "Price", "Quantity"), show="headings")
        self.tree_view.heading("ID", text="ID")
        self.tree_view.heading("Name", text="Name")
        self.tree_view.heading("Price", text="Price")
        self.tree_view.heading("Quantity", text="Quantity")
        self.tree_view.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_view.yview)
        self.tree_view.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_view.xview)
        self.tree_view.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Configure the grid layout for tree_frame to allow resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Refresh button
        ttk.Button(self.tab_view, text="Refresh", command=self.refresh_view_products).pack(pady=10)



    def refresh_view_products(self):
        for row in self.tree_view.get_children():
            self.tree_view.delete(row)
        for row in display_products():
            self.tree_view.insert("", "end", values=row)

################################################################################
##################purhcase product tab and its functionality########################

    def create_purchase_product_tab(self):
        # Create main frames to organize content
        input_frame = ttk.Frame(self.tab_purchase)
        cart_frame = ttk.Frame(self.tab_purchase)

        input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        cart_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Product Information Section
        ttk.Label(input_frame, text="Product Information", font=("Helvetica", 12, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(0, 10))

        labels = ["Product ID", "Quantity", "Discount (%)", "GST (%)"]
        entries = ["entry_purchase_product_id", "entry_purchase_quantity", "entry_discount", "entry_gst"]

        for i, (label, entry) in enumerate(zip(labels, entries)):
            ttk.Label(input_frame, text=label).grid(row=i + 1, column=0, padx=5, pady=5, sticky="e")
            setattr(self, entry, ttk.Entry(input_frame))
            getattr(self, entry).grid(row=i + 1, column=1, padx=5, pady=5, sticky="ew")

        # Customer Information Section
        ttk.Label(input_frame, text="Customer Information", font=("Helvetica", 12, "bold")).grid(
            row=5, column=0, columnspan=2, pady=(20, 10))

        # Customer details
        customer_info = [
            ("Customer Name", "entry_customer_name"),
            ("Customer Phone", "entry_customer_phone"),
            ("Payment Type", "payment_type_combo"),
            ("Amount Paid", "entry_amount_paid")
        ]

        for i, (label, entry) in enumerate(customer_info):
            ttk.Label(input_frame, text=label).grid(row=i + 6, column=0, padx=5, pady=5, sticky="e")
            if label == "Payment Type":
                self.payment_type = tk.StringVar(value="Cash")
                self.payment_type_combo = ttk.Combobox(input_frame, textvariable=self.payment_type)
                self.payment_type_combo['values'] = ('Cash', 'Credit')
                self.payment_type_combo.grid(row=i + 6, column=1, padx=5, pady=5, sticky="ew")
            else:
                setattr(self, entry, ttk.Entry(input_frame))
                getattr(self, entry).grid(row=i + 6, column=1, padx=5, pady=5, sticky="ew")

        # Buttons
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=10, column=0, columnspan=2, pady=20)

        self.add_to_cart_button = ttk.Button(button_frame, text="Add to Cart", command=self.add_to_cart)
        self.add_to_cart_button.grid(row=0, column=0, padx=5)

        self.view_cart_button = ttk.Button(button_frame, text="View Cart", command=self.view_cart)
        self.view_cart_button.grid(row=0, column=1, padx=5)

        # Cart Section (right side)
        ttk.Label(cart_frame, text="Shopping Cart", font=("Helvetica", 12, "bold")).grid(
            row=0, column=0, pady=(0, 10))

        # Treeview for cart items
        self.cart_tree = []
        self.cart_total = 0

        tree_container = ttk.Frame(cart_frame)
        tree_container.grid(row=1, column=0, sticky="nsew")

        self.cart_tree_view = ttk.Treeview(
            tree_container,
            columns=("ID", "Name", "Price", "Quantity", "Base Total", "Discount", "GST", "Final Total"),
            show="headings"
        )

        # Configure column headings
        for col in self.cart_tree_view["columns"]:
            self.cart_tree_view.heading(col, text=col)
            self.cart_tree_view.column(col, width=100)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.cart_tree_view.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.cart_tree_view.xview)
        self.cart_tree_view.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Grid layout for treeview and scrollbars
        self.cart_tree_view.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Total section
        total_frame = ttk.Frame(cart_frame)
        total_frame.grid(row=2, column=0, pady=10, sticky="e")

        ttk.Label(total_frame, text="Total:").grid(row=0, column=0, padx=5)
        self.cart_total_label = ttk.Label(total_frame, text="0.00")
        self.cart_total_label.grid(row=0, column=1, padx=5)

        # Configure weights for proper resizing
        self.tab_purchase.grid_columnconfigure(1, weight=3)
        self.tab_purchase.grid_columnconfigure(0, weight=1)
        self.tab_purchase.grid_rowconfigure(0, weight=1)

        cart_frame.grid_columnconfigure(0, weight=1)
        cart_frame.grid_rowconfigure(1, weight=1)

        tree_container.grid_columnconfigure(0, weight=1)
        tree_container.grid_rowconfigure(0, weight=1)

    def add_to_cart(self):
        product_id = int(self.entry_purchase_product_id.get())
        quantity = int(self.entry_purchase_quantity.get())
        # Remove purchase_product call here since it's reducing stock immediately
        name = self.get_product_name(product_id)
        price = self.get_product_price(product_id)

        # Check if stock is available without updating it
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                current_stock = row[3].value
                if current_stock >= quantity:
                    # Calculate base price and totals with GST and discount
                    base_total = price * quantity
                    discount = float(self.entry_discount.get() if self.entry_discount.get() else 0)
                    gst = float(self.entry_gst.get() if self.entry_gst.get() else 0)

                    discount_amount = base_total * (discount / 100)
                    after_discount = base_total - discount_amount
                    gst_amount = after_discount * (gst / 100)
                    final_total = after_discount + gst_amount

                    # Store all values in cart_tree
                    self.cart_tree.append((product_id, name, price, quantity, base_total, discount, gst, final_total))
                    self.cart_total += final_total
                    self.refresh_cart()
                    messagebox.showinfo("Success", f"Added {quantity} items to cart. Available stock: {current_stock}")
                else:
                    messagebox.showerror("Error", f"Not enough stock. Available stock: {current_stock}")
                break
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def update_stock_quantity(self, product_id, quantity):
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[3].value -= quantity
                wb.save(WORKBOOK_PATH)
                return True
        return False

    def revert_quantity(self,  product_id, quantity):
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[3].value += quantity
                wb.save(WORKBOOK_PATH)
                return True
        return False

    def clear_cart(self):
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            self.revert_quantity( product_id, quantity)
        self.cart_tree = []
        self.cart_total = 0
        self.refresh_cart()
        self.cart_window.destroy()

    def view_cart(self):
        self.create_cart_window()

    def create_cart_window(self):
        self.cart_window = tk.Toplevel(self.root)
        self.cart_window.title("Shopping Cart")
        self.cart_window.geometry("800x700")
        self.cart_window.configure(bg='#34495E')

        # For cart window
        self.cart_window.resizable(True, True)
        self.cart_window.minsize(820, 700)

        # Header
        header_frame = ttk.Frame(self.cart_window)
        header_frame.pack(fill='x', pady=10)
        ttk.Label(header_frame, text="Your Shopping Cart",
                  font=('Helvetica', 16, 'bold')).pack()

        # Cart content
        cart_frame = ttk.Frame(self.cart_window)
        cart_frame.pack(padx=20, pady=10, fill='both', expand=True)

        # Styled text widget
        cart_text = tk.Text(cart_frame, height=20, width=110,
                            font=('Courier', 10),
                            bg='white',
                            relief='ridge',
                            borderwidth=2)
        cart_text.pack(padx=10, pady=10)

        # Header formatting with colors - removed extra newline and spaces
        cart_data = f"{'ID':<10}{'Name':<20}{'Price':<10}{'Quantity':<10}{'Base Total':<12}{'Discount':<10}{'GST':<10}{'Final Total':<10}\n"
        cart_text.insert(tk.END, cart_data)
        cart_text.tag_add("header", "1.0", "2.0")  # Changed to start from first line
        cart_text.tag_config("header", background="#4a90e2", foreground="white", font=('Courier', 10, 'bold'))

        # Cart items
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            item_data = f"{product_id:<10}{name:<20}{price:<10.2f}{quantity:<10}{base_total:<12.2f}{discount:<10}%{gst:<10}%{final_total:<10.2f}\n"
            cart_text.insert(tk.END, item_data)

        # Total amount with styling
        total_text = f"\nTotal Amount: /-{self.cart_total:.2f}"
        cart_text.insert(tk.END, total_text)
        cart_text.tag_add("total", f"end-2c linestart", "end")
        cart_text.tag_config("total", font=('Helvetica', 12, 'bold'))

        # Buttons frame
        button_frame = ttk.Frame(self.cart_window)
        button_frame.pack(pady=20, padx=20)

        # Style configuration for buttons
        style = ttk.Style()
        style.configure('Action.TButton', padding=10, font=('Helvetica', 10))
        # Style configuration for buttons
        # style = ttk.Style()

        # Define custom styles for each button
        style.configure('Clear.TButton', padding=10, font=('Helvetica', 10), background='#e74c3c')  # Soft red
        style.configure('Checkout.TButton', padding=10, font=('Helvetica', 10), background='#2ecc71')  # Soft green
        style.configure('Generate.TButton', padding=10, font=('Helvetica', 10), background='#3498db')  # Soft blue
        style.configure('Print.TButton', padding=10, font=('Helvetica', 10), background='#9b59b6')  # Soft purple

        # Buttons with improved layout and colors
        buttons = [
            ("Clear Cart", self.clear_cart, 'Clear.TButton'),
            ("Checkout", self.checkout, 'Checkout.TButton'),
            ("Generate Bill", self.generate_bill, 'Generate.TButton'),
            ("Print Bill", self.print_bill, 'Print.TButton')
        ]

        for text, command, style_name in buttons:
            btn = ttk.Button(button_frame, text=text, command=command, style=style_name)
            btn.pack(side='left', padx=10)

        # Make text widget read-only
        cart_text.configure(state='disabled')

    def refresh_cart(self):
        for row in self.cart_tree_view.get_children():
            self.cart_tree_view.delete(row)
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            self.cart_tree_view.insert("", "end", values=(product_id, name, price, quantity, base_total,
                                                          f"{discount}%", f"{gst}%", final_total))
        self.cart_total_label.config(text=f"{self.cart_total:.2f}")

    def checkout(self):
        if not self.cart_tree:
            messagebox.showerror("Error", "Cart is empty")
            return

        # Now update stock quantities only during checkout
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            self.update_stock_quantity(product_id, quantity)

        self.generate_bill()
        self.record_sales()
        self.cart_tree = []
        self.cart_total = 0
        self.refresh_cart()
        self.cart_window.destroy()

    def record_sales(self):
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            customer_name = self.entry_customer_name.get()
            customer_phone = self.entry_customer_phone.get()
            payment_type = self.payment_type.get()
            amount_paid = float(self.entry_amount_paid.get() or 0)
            amount_due = final_total - amount_paid if payment_type == "Credit" else 0

            sale_data = [
                date, product_id, name, quantity, price, base_total,
                discount, gst, final_total,
                customer_name, customer_phone,
                payment_type, amount_paid, amount_due
            ]
            record_sale(sale_data)

    def generate_bill(self):
        self.bill_window = tk.Toplevel(self.root)
        self.bill_window.title("Bill Receipt")
        self.bill_window.configure(bg='#f0f0f0')

        # Set window size and center it
        window_width = 600
        window_height = 700
        screen_width = self.bill_window.winfo_screenwidth()
        screen_height = self.bill_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.bill_window.geometry(f'{window_width}x{window_height}+{x}+{y}')

        # Create main frame
        main_frame = ttk.Frame(self.bill_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Configure Text widget with better styling
        self.bill_text = tk.Text(
            main_frame,
            height=30,
            width=80,
            font=('Courier', 10),
            bg='white',
            relief='solid'
        )
        self.bill_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Create the bill content with better formatting
        bill_data = f"""
    ╔══════════════════════════════════════════════════════════════╗
                            SS INTERPRISORS
                      Address: I-10/MARKAZ, ISLAMABAD
                        Contact: +92 333 5130796
    ╚══════════════════════════════════════════════════════════════╝

    Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    Customer Name: {self.entry_customer_name.get()}
    Customer Phone: {self.entry_customer_phone.get()}
    Payment Type: {self.payment_type.get()}

    {'═' * 60}
       PRODUCT DETAILS
    {'═' * 60}
    """

        # Add items with better alignment
        discount=0
        gst=0
        subtotal = 0
        for item in self.cart_tree:
            product_id, name, price, quantity, base_total, discount, gst, final_total = item
            bill_data += f"│ {name:<20} │ Price: {price:>8} │ Qty: {quantity:>3} │ Total: {final_total:>8.2f} │\n"
            subtotal += base_total
            discount=discount
            gst=gst

        discount_amount = subtotal * (discount / 100)
        amount_after_discount = subtotal - discount_amount
        gst_amount = amount_after_discount * (gst / 100)
        final_total = amount_after_discount + gst_amount
        amount_paid = float(self.entry_amount_paid.get() or 0)
        amount_due = final_total - amount_paid

        # Add summary with better formatting
        bill_data += f"""
    {'═' * 60}
                            BILL SUMMARY
    {'═' * 60}
        Subtotal:                               {subtotal:>10.2f}
        Discount ({discount}%):                 {discount_amount:>10.2f}
        GST ({gst}%):                           {gst_amount:>10.2f}
        ────────────────────────────────────────────────────
        Final Total:                            {final_total:>10.2f}
        Amount Paid:                            {amount_paid:>10.2f}
        Amount Due:                             {amount_due:>10.2f}
    {'═' * 60}

                Thank you for your business!
                Please visit again.
    """

        self.bill_text.insert(tk.END, bill_data)

        # Create a styled print button
        style = ttk.Style()
        style.configure('Print.TButton', font=('Arial', 10, 'bold'))
        print_btn = ttk.Button(
            main_frame,
            text="Print Bill",
            style='Print.TButton',
            command=self.print_bill
        )
        print_btn.pack(pady=10)


    def print_bill(self):
        bill_text = self.bill_text.get(1.0, tk.END)
        print(bill_text)
        # Optionally, you could implement printing to a physical printer using external libraries or OS-specific commands.

    def get_product_name(self, product_id):
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["products"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                return row[1]
        return ""

    def get_product_price(self, product_id):
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["products"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                return row[2]
        return 0

################################################################################
#################################admin tab######################################
    def create_admin_tab(self):
        self.admin_notebook = ttk.Notebook(self.tab_admin)
        self.admin_notebook.pack(pady=20, expand=True, fill=tk.BOTH)

        self.tab_search = ttk.Frame(self.admin_notebook)
        self.tab_update = ttk.Frame(self.admin_notebook)
        self.tab_delete = ttk.Frame(self.admin_notebook)
        self.tab_order = ttk.Frame(self.admin_notebook)
        self.tab_sales_report = ttk.Frame(self.admin_notebook)  # New Sales Report tab
        self.tab_user_management = ttk.Frame(self.admin_notebook) # User Management Tab
              # Create a notebook within the User Management tab
        self.user_management_notebook = ttk.Notebook(self.tab_user_management)
        self.user_management_notebook.pack(expand=True, fill=tk.BOTH)
        self.admin_notebook.add(self.tab_user_management, text="User Management")
        self.admin_notebook.add(self.tab_search, text="Search Product")
        self.admin_notebook.add(self.tab_update, text="Update Product")
        self.admin_notebook.add(self.tab_delete, text="Delete Product")
        self.admin_notebook.add(self.tab_order, text="Place Order")
        self.admin_notebook.add(self.tab_sales_report, text="Sales Report")  # Adding Sales Report tab
        self.tab_add_user = ttk.Frame(self.user_management_notebook)
        self.tab_delete_user = ttk.Frame(self.user_management_notebook)
        self.tab_update_user = ttk.Frame(self.user_management_notebook)
        self.tab_show_all_users = ttk.Frame(self.user_management_notebook)

        # Adding sub-tabs to the User Management notebook
        self.user_management_notebook.add(self.tab_add_user, text="Add User")
        self.user_management_notebook.add(self.tab_delete_user, text="Delete User")
        self.user_management_notebook.add(self.tab_update_user, text="Update User")
        self.user_management_notebook.add(self.tab_show_all_users, text="Show All Users")
        self.create_search_product_tab()
        self.create_update_product_tab()
        self.create_delete_product_tab()
        self.create_order_product_tab()
        self.create_sales_report_tab()  # Call to create the sales report tab
        # You can now define functions to populate these tabs with relevant widgets and functionality
        self.create_add_user_tab()
        self.create_delete_user_tab()
        self.create_update_user_tab()
        self.create_show_all_users_tab()
####################################################################
########################add user tab and its functionality##################
    def create_add_user_tab(self):
        # Add widgets for adding a user
        row = 0
        ttk.Label(self.tab_add_user, text="Full Name:").grid(row=row, column=0, padx=5, pady=5)
        self.name_entry = ttk.Entry(self.tab_add_user)
        self.name_entry.grid(row=row, column=1, padx=5, pady=5)

        row += 1
        ttk.Label(self.tab_add_user, text="Contact Number:").grid(row=row, column=0, padx=5, pady=5)
        self.contact_entry = ttk.Entry(self.tab_add_user)
        self.contact_entry.grid(row=row, column=1, padx=5, pady=5)

        row += 1
        ttk.Label(self.tab_add_user, text="Username:").grid(row=row, column=0, padx=5, pady=5)
        self.username_entry = ttk.Entry(self.tab_add_user)
        self.username_entry.grid(row=row, column=1, padx=5, pady=5)

        row += 1
        ttk.Label(self.tab_add_user, text="User Type:").grid(row=row, column=0, padx=5, pady=5)
        self.user_type = ttk.Combobox(self.tab_add_user, values=['Admin', 'Manager', 'Cashier'], state='readonly')
        self.user_type.grid(row=row, column=1, padx=5, pady=5)

        row += 1
        ttk.Label(self.tab_add_user, text="Password:").grid(row=row, column=0, padx=5, pady=5)
        self.password_entry = ttk.Entry(self.tab_add_user, show="*")
        self.password_entry.grid(row=row, column=1, padx=5, pady=5)

        row += 1
        self.add_button = ttk.Button(self.tab_add_user, text="Add User", command=self.add_user)
        self.add_button.grid(row=row, column=0, columnspan=2, pady=20)
    def add_user(self):
        # Get values from entry fields
        full_name = self.name_entry.get()
        contact = self.contact_entry.get()
        username = self.username_entry.get()
        user_type = self.user_type.get()
        password = self.password_entry.get()

        # Basic validation
        if not all([full_name, contact, username, user_type, password]):
            messagebox.showerror("Error", "All fields are required")
            return

        try:
            # Load existing workbook
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["users"]

            # Check if username already exists
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == username:  # Username is in third column
                    messagebox.showerror("Error", "Username already exists")
                    return

            # Add new user
            ws.append([full_name, contact, username, user_type, password])
            wb.save(WORKBOOK_PATH)

            messagebox.showinfo("Success", "User added successfully")

            # Clear entry fields
            self.name_entry.delete(0, tk.END)
            self.contact_entry.delete(0, tk.END)
            self.username_entry.delete(0, tk.END)
            self.user_type.set('')
            self.password_entry.delete(0, tk.END)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
#######################################################################################
###################update tab and its fubctions######################
    def create_update_user_tab(self):
        # Create frames for search and update sections
        search_frame = ttk.Frame(self.tab_update_user)
        search_frame.pack(fill='x', padx=10, pady=5)

        update_frame = ttk.Frame(self.tab_update_user)
        update_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Search section
        ttk.Label(search_frame, text="Search Username:").grid(row=0, column=0, padx=5, pady=5)
        self.update_search_entry = ttk.Entry(search_frame)
        self.update_search_entry.grid(row=0, column=1, padx=5, pady=5)
        # self.update_search_entry.bind("<FocusOut>", lambda event: self.update_search_entry.update())

        ttk.Button(search_frame, text="Search", command=self.search_user_to_update).grid(row=0, column=2, padx=5,
                                                                                         pady=5)

        # Update section
        labels = ["Full Name:", "Contact:", "Username:", "User Type:", "New Password:"]
        entries = ["update_fullname_entry", "update_contact_entry", "update_username_entry",
                   "update_user_type_combo", "update_password_entry"]

        for i, (label, entry) in enumerate(zip(labels, entries)):
            ttk.Label(update_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')

            if label == "User Type:":
                setattr(self, entry,
                        ttk.Combobox(update_frame, values=['Admin', 'Manager', 'Cashier'], state='readonly'))
            elif label == "New Password:":
                setattr(self, entry, ttk.Entry(update_frame, show="*"))
            else:
                setattr(self, entry, ttk.Entry(update_frame))

            getattr(self, entry).grid(row=i, column=1, padx=5, pady=5, sticky='w')

            # Disable entries initially until user is searched
            getattr(self, entry).configure(state='disabled')

        # Update button
        self.update_user_button = ttk.Button(update_frame, text="Update User", command=self.update_user)
        self.update_user_button.grid(row=len(labels), column=0, columnspan=2, pady=20)
        self.update_user_button.configure(state='disabled')

        # Add these helper functions to the class
    def search_user_to_update(self):
        username = self.update_search_entry.get()

        if not username:
            messagebox.showerror("Error", "Please enter a username to search")
            return

        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["users"]

            user_found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == username:  # Username is in third column
                        # Enable all entry fields
                    self.update_fullname_entry.configure(state='normal')
                    self.update_contact_entry.configure(state='normal')
                    self.update_username_entry.configure(state='normal')
                    self.update_user_type_combo.configure(state='readonly')
                    self.update_password_entry.configure(state='normal')
                    self.update_user_button.configure(state='normal')

                        # Fill in current values
                    self.update_fullname_entry.delete(0, tk.END)
                    self.update_fullname_entry.insert(0, row[0])
                    self.update_contact_entry.delete(0, tk.END)
                    self.update_contact_entry.insert(0, row[1])
                    self.update_username_entry.delete(0, tk.END)
                    self.update_username_entry.insert(0, row[2])
                    self.update_user_type_combo.set(row[3])

                    user_found = True
                    break

            if not user_found:
                messagebox.showerror("Error", "User not found")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def update_user(self):
        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["users"]

            old_username = self.update_search_entry.get()
            new_name = self.update_fullname_entry.get()
            new_contact = self.update_contact_entry.get()
            new_username = self.update_username_entry.get()
            new_user_type = self.update_user_type_combo.get()
            new_password = self.update_password_entry.get()

                # Basic validation
            if not all([new_name, new_contact, new_username, new_user_type]):
                messagebox.showerror("Error", "All fields except password are required")
                return

            for row in ws.iter_rows(min_row=2):
                if row[2].value == old_username:
                    row[0].value = new_name
                    row[1].value = new_contact
                    row[2].value = new_username
                    row[3].value = new_user_type
                    if new_password:  # Update password only if a new one is provided
                        row[4].value = new_password
                    break

            wb.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "User updated successfully")

                # Clear and disable fields
            for entry in [self.update_fullname_entry, self.update_contact_entry,
                          self.update_username_entry, self.update_password_entry]:
                entry.delete(0, tk.END)
                entry.configure(state='disabled')

            self.update_user_type_combo.set('')
            self.update_user_type_combo.configure(state='disabled')
            self.update_user_button.configure(state='disabled')
            self.update_search_entry.delete(0, tk.END)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
#############################################################################################
#############################show users tab and functions#####################################
    def create_show_all_users_tab(self):
        # Create frame for treeview and scrollbars first
        tree_frame = ttk.Frame(self.tab_show_all_users)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Create a treeview to display all users
        columns = ('Full Name', 'Contact', 'Username', 'User Type')
        self.users_tree = ttk.Treeview(tree_frame, columns=columns, show='headings')

        # Set column headings
        for col in columns:
            self.users_tree.heading(col, text=col)
            self.users_tree.column(col, width=150)

        # Add scrollbars
        y_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=self.users_tree.yview)
        x_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.users_tree.xview)

        # Configure scrollbars
        self.users_tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)

        # Grid layout for treeview and scrollbars
        self.users_tree.grid(row=0, column=0, sticky='nsew')
        y_scrollbar.grid(row=0, column=1, sticky='ns')
        x_scrollbar.grid(row=1, column=0, sticky='ew')

        # Configure grid weights
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        # Create frame for controls last
        control_frame = ttk.Frame(self.tab_show_all_users)
        control_frame.pack(fill='x', padx=5, pady=5)

        # Add refresh button
        refresh_btn = ttk.Button(control_frame, text="Refresh", command=self.refresh_users_list)
        refresh_btn.pack(padx=5, pady=5)

        # Load initial data
        self.refresh_users_list()

    def refresh_users_list(self):
        # Clear existing items
        for item in self.users_tree.get_children():
            self.users_tree.delete(item)

        try:
            # Read the Excel file
            from openpyxl import load_workbook
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["users"]

            # Skip the header row and insert data into treeview
            for row in list(ws.rows)[1:]:
                values = [cell.value for cell in row]
                self.users_tree.insert('', 'end', values=values)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to")
################################################################################
####################dlete user tab and its functionality########################
    def create_delete_user_tab(self):
        # Create a frame for the search section
        search_frame = ttk.Frame(self.tab_delete_user)
        search_frame.pack(fill='x', padx=10, pady=5)

        # Search section
        ttk.Label(search_frame, text="Search Username:").grid(row=0, column=0, padx=5, pady=5)
        self.delete_search_entry = ttk.Entry(search_frame)
        self.delete_search_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(search_frame, text="Search", command=self.search_user_to_delete).grid(row=0, column=2, padx=5,
                                                                                         pady=5)

        # Create a frame for the delete section
        delete_frame = ttk.Frame(self.tab_delete_user)
        delete_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # User details section
        labels = ["Full Name:", "Contact:", "Username:", "User Type:"]
        entries = ["delete_name_entry", "delete_contact_entry", "delete_username_entry", "delete_user_type_combo"]

        for i, (label, entry) in enumerate(zip(labels, entries)):
            ttk.Label(delete_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')

            if label == "User Type:":
                setattr(self, entry,
                        ttk.Combobox(delete_frame, values=['Admin', 'Manager', 'Cashier'], state='readonly'))
            else:
                setattr(self, entry, ttk.Entry(delete_frame))

            getattr(self, entry).grid(row=i, column=1, padx=5, pady=5, sticky='w')

            # Disable entries initially until user is searched
            getattr(self, entry).configure(state='disabled')

        # Delete button
        self.delete_user_button = ttk.Button(delete_frame, text="Delete User", command=self.delete_user)
        self.delete_user_button.grid(row=len(labels), column=0, columnspan=2, pady=20)
        self.delete_user_button.configure(state='disabled')

    def search_user_to_delete(self):
        username = self.delete_search_entry.get()

        if not username:
            messagebox.showerror("Error", "Please enter a username to search")
            return

        try:
            wb = openpyxl.load_workbook("users.xlsx")
            ws = wb.active

            user_found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == username:  # Username is in third column
                    # Enable all entry fields
                    self.delete_name_entry.configure(state='normal')
                    self.delete_contact_entry.configure(state='normal')
                    self.delete_username_entry.configure(state='normal')
                    self.delete_user_type_combo.configure(state='readonly')
                    self.delete_user_button.configure(state='normal')

                    # Fill in current values
                    self.delete_name_entry.delete(0, tk.END)
                    self.delete_name_entry.insert(0, row[0])
                    self.delete_contact_entry.delete(0, tk.END)
                    self.delete_contact_entry.insert(0, row[1])
                    self.delete_username_entry.delete(0, tk.END)
                    self.delete_username_entry.insert(0, row[2])
                    self.delete_user_type_combo.set(row[3])

                    user_found = True
                    break

            if not user_found:
                messagebox.showerror("Error", "User not found")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def delete_user(self):
        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["users"]

            username = self.delete_search_entry.get()

            for row in ws.iter_rows(min_row=2):
                if row[2].value == username:  # Username is in third column
                    ws.delete_rows(row[0].row, 1)
                    wb.save(WORKBOOK_PATH)
                    messagebox.showinfo("Success", "User deleted successfully")

                    # Clear and disable fields
                    for entry in [self.delete_name_entry, self.delete_contact_entry, self.delete_username_entry]:
                        entry.delete(0, tk.END)
                        entry.configure(state='disabled')

                    self.delete_user_type_combo.set('')
                    self.delete_user_type_combo.configure(state='disabled')
                    self.delete_user_button.configure(state='disabled')
                    self.delete_search_entry.delete(0, tk.END)
                    break

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


###############################################################################
##########################sales report tab and its functionality###############
    def create_sales_report_tab(self):
        # Create a top frame for controls
        top_frame = ttk.Frame(self.tab_sales_report)
        top_frame.grid(row=0, column=0, columnspan=5, padx=10, pady=5, sticky="EW")

        # First row: Financial Summary
        ttk.Label(top_frame, text="Total Sales:", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=10, pady=5,
                                                                                   sticky="W")
        self.total_sales_var = tk.StringVar()
        ttk.Label(top_frame, textvariable=self.total_sales_var, font=("Arial", 10)).grid(row=0, column=1, padx=10,
                                                                                         pady=5, sticky="W")
        ttk.Label(top_frame, text="PAID Amount:", font=("Arial", 10, "bold")).grid(row=0, column=2, padx=10, pady=5,
                                                                                      sticky="W")
        self.paid_amount_var = tk.StringVar()
        ttk.Label(top_frame, textvariable=self.paid_amount_var, font=("Arial", 10)).grid(row=0, column=3, padx=10,
                                                                                            pady=5, sticky="W")
        ttk.Label(top_frame, text="Credit Balance:", font=("Arial", 10, "bold")).grid(row=0, column=4, padx=10, pady=5,
                                                                                      sticky="W")
        self.credit_balance_var = tk.StringVar()
        ttk.Label(top_frame, textvariable=self.credit_balance_var, font=("Arial", 10)).grid(row=0, column=5, padx=10,
                                                                                            pady=5, sticky="W")

        # Second row: Search
        ttk.Label(top_frame, text="Search:", font=("Arial", 10, "bold")).grid(row=1, column=0, padx=10, pady=5,
                                                                              sticky="W")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(top_frame, textvariable=self.search_var, width=20)
        search_entry.grid(row=1, column=1, padx=10, pady=5, sticky="W")
        search_button = ttk.Button(top_frame, text="Search", command=self.search_sales_record)
        search_button.grid(row=1, column=2, padx=10, pady=5, sticky="W")

        # Third row: Filter Controls
        ttk.Label(top_frame, text="Select Month", font=("Arial", 10, "bold")).grid(row=2, column=0, padx=10, pady=5,
                                                                                   sticky="W")
        self.month_var = tk.StringVar()
        self.month_dropdown = ttk.Combobox(top_frame, textvariable=self.month_var, width=15)
        self.month_dropdown['values'] = [month_name[i] for i in range(1, 13)]
        self.month_dropdown.grid(row=2, column=1, padx=10, pady=5, sticky="W")

        ttk.Label(top_frame, text="Select Year", font=("Arial", 10, "bold")).grid(row=2, column=2, padx=10, pady=5,
                                                                                  sticky="W")
        self.year_var = tk.StringVar()
        self.year_dropdown = ttk.Combobox(top_frame, textvariable=self.year_var, width=15)
        self.year_dropdown['values'] = [str(year) for year in range(2000, datetime.now().year + 1)]
        self.year_dropdown.grid(row=2, column=3, padx=10, pady=5, sticky="W")

        filter_button = ttk.Button(top_frame, text="Filter", command=self.filter_sales_report)
        filter_button.grid(row=2, column=4, padx=10, pady=5, sticky="W")

        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_sales_report)
        tree_frame.grid(row=1, column=0, columnspan=5, padx=10, pady=10, sticky="NSEW")

        # Treeview setup (rest of the Treeview code remains the same)
        self.tree_sales = ttk.Treeview(tree_frame, columns=(
            "Date", "Product ID", "Product Name", "Quantity", "Price", "Total",
            "Discount", "GST", "Final Total", "Customer Name", "Customer Phone",
            "Payment Type", "Amount Paid", "Amount Due"), show="headings")

        # Set column headings and widths
        for col in self.tree_sales["columns"]:
            self.tree_sales.heading(col, text=col)
            if col in ["Discount", "GST", "Price", "Total", "Final Total", "Amount Paid", "Amount Due"]:
                self.tree_sales.column(col, width=100, anchor="e")  # Right-align numeric columns
            elif col in ["Date", "Product ID"]:
                self.tree_sales.column(col, width=120)
            else:
                self.tree_sales.column(col, width=150)

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_sales.yview)
        self.tree_sales.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_sales.xview)
        self.tree_sales.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Position the Treeview in the frame
        self.tree_sales.grid(row=0, column=0, sticky="nsew")

        # Configure tree_frame to be resizable
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        # Configure grid expansion for resizing
        self.tab_sales_report.columnconfigure(4, weight=1)
        self.tab_sales_report.rowconfigure(1, weight=1)

        # Load all sales initially
        self.load_sales_data()

        # Update the total sales and credit balance initially
        self.update_financials()

    def search_sales_record(self):
        search_text = self.search_var.get().lower()
        # First, hide all records
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)

        # Load and filter data from the Excel file
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["sales"]

        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Convert row to list for formatting
            formatted_row = list(row)
            # Format numeric values
            for i, value in enumerate(formatted_row):
                if i in [4, 5, 6, 7, 8, 12, 13]:  # Numeric columns
                    try:
                        formatted_row[i] = float(value) if value is not None else 0.0
                    except (ValueError, TypeError):
                        formatted_row[i] = 0.0

            # Check if search text exists in any field
            if search_text in str(formatted_row).lower():
                self.tree_sales.insert("", "end", values=formatted_row)
                found = True

        if not found:
            messagebox.showinfo("Search", "No matching record found.")
            self.load_sales_data()  # Reload all data if no matches found

        # Update financial summary for displayed records
        self.update_financials()

    def update_financials(self):
        total_sales = 0.0
        paid_amount=0.0
        credit_balance = 0.0
        for row in self.tree_sales.get_children():
            row_data = self.tree_sales.item(row)['values']
            # Convert string to float for Final Total
            total_sales += float(row_data[8] if row_data[8] else 0)
            paid_amount+=float(row_data[12] if row_data[12] else 0)
            # Convert string to float for Amount Due
            amount_due = float(row_data[13] if row_data[13] else 0)
            if amount_due > 0:
                credit_balance += amount_due

        self.total_sales_var.set(f"/-{total_sales:,.2f}")
        self.paid_amount_var.set(f"/-{paid_amount:,.2f}")
        self.credit_balance_var.set(f"/-{credit_balance:,.2f}")

    def load_sales_data(self):
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)

        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["sales"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Convert numeric values to float and format them
            formatted_row = list(row)
            for i, value in enumerate(row):
                # Format numeric columns (Price, Total, Discount, GST, etc.)
                if i in [4, 5, 6, 7, 8, 12, 13]:  # Indices for numeric columns
                    try:
                        formatted_row[i] = float(value) if value is not None else 0.0
                    except (ValueError, TypeError):
                        formatted_row[i] = 0.0

            self.tree_sales.insert("", "end", values=formatted_row)

    def filter_sales_report(self):
        """Filters sales based on selected month and year"""
        selected_month = self.month_var.get()
        selected_year = self.year_var.get()

        if not selected_month or not selected_year:
            messagebox.showerror("Error", "Please select both month and year.")
            return

        month_num = list(month_name).index(selected_month)  # Convert month name to month number
        filtered_sales = []

        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["sales"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sale_date = row[0]  # Get the sale date

            # Check if sale_date is a datetime object
            if isinstance(sale_date, datetime):
                if sale_date.month == month_num and sale_date.year == int(selected_year):
                    filtered_sales.append(row)
            else:
                # If it's a string, parse it
                sale_date = datetime.strptime(sale_date, "%Y-%m-%d %H:%M:%S")
                if sale_date.month == month_num and sale_date.year == int(selected_year):
                    filtered_sales.append(row)

        # Clear the tree view and insert filtered data
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)
        for row in filtered_sales:
            self.tree_sales.insert("", "end", values=row)

    #######################################################################################
###############################search product tab and its functionality################
    def create_search_product_tab(self):
        ttk.Label(self.tab_search, text="Search Term").grid(row=0, column=0, padx=10, pady=10)
        self.entry_search_term = ttk.Entry(self.tab_search)
        self.entry_search_term.grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(self.tab_search, text="Search", command=self.search_product).grid(row=1, column=0, columnspan=2,
                                                                                     pady=10)

        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_search)
        tree_frame.grid(row=2, column=0, columnspan=2, pady=10, padx=10, sticky="nsew")

        # Treeview for displaying search results
        self.tree_search = ttk.Treeview(tree_frame, columns=("ID", "Name", "Price", "Quantity"), show="headings")
        self.tree_search.heading("ID", text="ID")
        self.tree_search.heading("Name", text="Name")
        self.tree_search.heading("Price", text="Price")
        self.tree_search.heading("Quantity", text="Quantity")
        self.tree_search.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_search.yview)
        self.tree_search.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_search.xview)
        self.tree_search.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Configure the grid layout for tree_frame to allow resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure the grid expansion for the main tab to allow resizing
        self.tab_search.columnconfigure(1, weight=1)
        self.tab_search.rowconfigure(2, weight=1)

    def search_product(self):
        search_term = self.entry_search_term.get()
        for row in self.tree_search.get_children():
            self.tree_search.delete(row)
        for row in search_product( search_term):
            self.tree_search.insert("", "end", values=row)
######################################################################################
######################################update product tab and its functionality########
    def create_update_product_tab(self):
        # Create frames for search and update sections
        search_frame1 = ttk.Frame(self.tab_update)
        search_frame1.pack(fill='x', padx=10, pady=5)

        update_frame1 = ttk.Frame(self.tab_update)
        update_frame1.pack(fill='both', expand=True, padx=10, pady=5)

        # Search section
        ttk.Label(search_frame1, text="Search Product ID:").grid(row=0, column=0, padx=5, pady=5)
        self.update_product_id_entry = ttk.Entry(search_frame1)
        self.update_product_id_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(search_frame1, text="Search", command=self.search_product_to_update).grid(row=0, column=2, padx=5,
                                                                                            pady=5)

        # Update section
        labels = ["Product ID:", "Name:", "Price:", "Quantity:"]
        entries = ["update_id_entry", "update_name_entry", "update_price_entry", "update_quantity_entry"]

        for i, (label, entry) in enumerate(zip(labels, entries)):
            ttk.Label(update_frame1, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            setattr(self, entry, ttk.Entry(update_frame1))
            getattr(self, entry).grid(row=i, column=1, padx=5, pady=5, sticky='w')
            getattr(self, entry).configure(state='disabled')

        # Update button
        self.update_product_button = ttk.Button(update_frame1, text="Update Product", command=self.update_product)
        self.update_product_button.grid(row=len(labels), column=0, columnspan=2, pady=20)
        self.update_product_button.configure(state='disabled')

    def search_product_to_update(self):
        product_id = self.update_product_id_entry.get()

        if not product_id:
            messagebox.showerror("Error", "Please enter a Product ID to search")
            return

        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["products"]

            product_found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == product_id:
                    # Enable all entry fields
                    self.update_id_entry.configure(state='normal')
                    self.update_name_entry.configure(state='normal')
                    self.update_price_entry.configure(state='normal')
                    self.update_quantity_entry.configure(state='normal')
                    self.update_product_button.configure(state='normal')

                    # Fill in current values
                    self.update_id_entry.delete(0, tk.END)
                    self.update_id_entry.insert(0, row[0])
                    self.update_name_entry.delete(0, tk.END)
                    self.update_name_entry.insert(0, row[1])
                    self.update_price_entry.delete(0, tk.END)
                    self.update_price_entry.insert(0, row[2])
                    self.update_quantity_entry.delete(0, tk.END)
                    self.update_quantity_entry.insert(0, row[3])

                    product_found = True
                    break

            if not product_found:
                messagebox.showerror("Error", "Product not found")
            else:
                messagebox.showinfo("Success", "Product loaded for update.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def update_product(self):
        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)

            ws = wb["products"]

            product_id = self.update_product_id_entry.get()
            new_id = self.update_id_entry.get()
            new_name = self.update_name_entry.get()
            new_price = self.update_price_entry.get()
            new_quantity = self.update_quantity_entry.get()

            # Basic validation
            if not all([new_id, new_name, new_price, new_quantity]):
                messagebox.showerror("Error", "All fields are required")
                return

            try:
                new_price = float(new_price)
                new_quantity = int(new_quantity)
            except ValueError:
                messagebox.showerror("Error", "Price must be a number, and quantity must be an integer")
                return

            product_updated = False
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == product_id:
                    row[0].value = int(new_id)
                    row[1].value = new_name
                    row[2].value = new_price
                    row[3].value = new_quantity
                    product_updated = True
                    break

            if product_updated:
                wb.save(WORKBOOK_PATH)
                messagebox.showinfo("Success", "Product updated successfully")
            else:
                messagebox.showerror("Error", "Product ID not found")

            # Clear fields after update
            for entry in [self.update_id_entry, self.update_name_entry,
                          self.update_price_entry, self.update_quantity_entry]:
                entry.delete(0, tk.END)
                entry.configure(state='disabled')

            self.update_product_button.configure(state='disabled')
            self.update_product_id_entry.delete(0, tk.END)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    ######################################################################################
######################################delete product tab and its functionality########

    def create_delete_product_tab(self):
        # Create frames for search and delete sections
        search_frame2 = ttk.Frame(self.tab_delete)
        search_frame2.pack(fill='x', padx=10, pady=5)

        delete_frame = ttk.Frame(self.tab_delete)
        delete_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Search section
        ttk.Label(search_frame2, text="Search Product ID:").grid(row=0, column=0, padx=5, pady=5)
        self.delete_product_id_entry = ttk.Entry(search_frame2)
        self.delete_product_id_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(search_frame2, text="Search", command=self.search_product_to_delete).grid(row=0, column=2, padx=5,
                                                                                            pady=5)

        # Product details section
        labels = ["Product ID:", "Name:", "Price:", "Quantity:"]
        entries = ["delete_id_entry", "delete_pname_entry", "delete_price_entry", "delete_quantity_entry"]

        for i, (label, entry) in enumerate(zip(labels, entries)):
            ttk.Label(delete_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            setattr(self, entry, ttk.Entry(delete_frame))
            getattr(self, entry).grid(row=i, column=1, padx=5, pady=5, sticky='w')
            getattr(self, entry).configure(state='disabled')

        # Delete button_product_button = ttk.Button(delete_frame, text="Delete Product", command=self.delete_product)
        self.delete_product_button = ttk.Button(delete_frame, text="Delete Product", command=self.delete_product)
        self.delete_product_button.grid(row=len(labels), column=1, pady=10)
        self.delete_product_button.configure(state='disabled')  # Initially disabled

        self.delete_product_button.grid(row=len(labels), column=0, columnspan=2, pady=20)
        self.delete_product_button.configure(state='disabled')

    def search_product_to_delete(self):
        product_id = self.delete_product_id_entry.get()

        if not product_id:
            messagebox.showerror("Error", "Please enter a Product ID to search")
            return

        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)

            ws = wb["products"]

            product_found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == product_id:
                    # Enable all entry fields
                    self.delete_id_entry.configure(state='normal')
                    self.delete_pname_entry.configure(state='normal')
                    self.delete_price_entry.configure(state='normal')
                    self.delete_quantity_entry.configure(state='normal')
                    self.delete_product_button.configure(state='normal')

                    # Fill in current values
                    self.delete_id_entry.delete(0, tk.END)
                    self.delete_id_entry.insert(0, row[0])
                    self.delete_pname_entry.delete(0, tk.END)
                    self.delete_pname_entry.insert(0, row[1])
                    self.delete_price_entry.delete(0, tk.END)
                    self.delete_price_entry.insert(0, row[2])
                    self.delete_quantity_entry.delete(0, tk.END)
                    self.delete_quantity_entry.insert(0, row[3])

                    product_found = True
                    break

            if not product_found:
                messagebox.showerror("Error", "Product not found")
            else:
                messagebox.showinfo("Success", "Product loaded for deletion.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def delete_product(self):
        try:
            wb = openpyxl.load_workbook(WORKBOOK_PATH)
            ws = wb["products"]

            product_id = self.delete_product_id_entry.get()

            confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this product?")
            if not confirm:
                return

            product_deleted = False
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == product_id:
                    ws.delete_rows(row[0].row, 1)
                    product_deleted = True
                    break

            if product_deleted:
                wb.save(WORKBOOK_PATH)
                messagebox.showinfo("Success", "Product deleted successfully")
            else:
                messagebox.showerror("Error", "Product ID not found")

            # Clear fields after deletion
            for entry in [self.delete_id_entry, self.delete_pname_entry,
                          self.delete_price_entry, self.delete_quantity_entry]:
                entry.delete(0, tk.END)
                entry.configure(state='disabled')

            self.delete_product_button.configure(state='disabled')
            self.delete_search_entry.delete(0, tk.END)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    ######################################################################################
######################################order product tab and its functionality########

    def create_order_product_tab(self):
        ttk.Label(self.tab_order, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_order_product_id = ttk.Entry(self.tab_order)
        self.entry_order_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_order, text="Order Quantity").grid(row=1, column=0, padx=10, pady=10)
        self.entry_order_quantity = ttk.Entry(self.tab_order)
        self.entry_order_quantity.grid(row=1, column=1, padx=10, pady=10)

        ttk.Button(self.tab_order, text="Place Order", command=self.place_order_product).grid(row=2, column=0, columnspan=2, pady=10)

    def place_order_product(self):
        product_id = int(self.entry_order_product_id.get())
        order_quantity = int(self.entry_order_quantity.get())
        if place_order_product(product_id, order_quantity):
            messagebox.showinfo("Success", "Order placed successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

######################################################################################
######################################bill window###############################

    def create_bill_window(self):
        self.bill_window = tk.Toplevel(self.root)
        self.bill_window.title("Bill Receipt")

        # For bill window
        self.bill_window.resizable(True, True)
        self.bill_window.minsize(600, 400)
        self.bill_text = tk.Text(self.bill_window, height=20, width=80)
        self.bill_text.pack(padx=20, pady=20)

        ttk.Button(self.bill_window, text="Print", command=self.print_bill).pack(pady=10)
        ttk.Button(self.bill_window, text="Generate Bill", command=self.generate_bill).pack(pady=10)

#####################################################################################################

    def clear_entries(self):
        self.entry_product_id.delete(0, tk.END)
        self.entry_product_name.delete(0, tk.END)
        self.entry_product_price.delete(0, tk.END)
        self.entry_product_quantity.delete(0, tk.END)
        self.entry_purchase_product_id.delete(0, tk.END)
        self.entry_purchase_quantity.delete(0, tk.END)
        self.entry_discount.delete(0, tk.END)
        self.entry_gst.delete(0, tk.END)
        self.entry_search_term.delete(0, tk.END)
        # self.entry_update_product_id.delete(0, tk.END)
        # self.entry_update_name.delete(0, tk.END)
        # self.entry_update_price.delete(0, tk.END)
        # self.entry_update_quantity.delete(0, tk.END)
        # self.entry_delete_product_id.delete(0, tk.END)
        self.entry_order_product_id.delete(0, tk.END)
        self.entry_order_quantity.delete(0, tk.END)

if __name__ == "__main__":
    initialize_workbook()
    root = tk.Tk()
    app = LoginWindow(root)
    root.mainloop()
