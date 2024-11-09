import tkinter as tk
from tkinter import ttk, messagebox, font
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
from calendar import month_name

# Ensure the Excel file is created
def create_product_excel(file_name="products.xlsx"):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["ID", "Name", "Price", "Quantity"])
        wb.save(file_name)

def create_sales_excel(file_name="sales.xlsx"):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Date", "Product ID", "Product Name", "Quantity", "Price", "Total", "Discount", "GST", "Final Total"])
        wb.save(file_name)

def record_sale(sale_data, file_name="sales.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.append(sale_data)
    wb.save(file_name)

class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Login")
        self.root.geometry("300x150")
        self.root.configure(bg="#2C3E50")

        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.root, text="Username").grid(row=0, column=0, padx=10, pady=10)
        self.entry_username = ttk.Entry(self.root)
        self.entry_username.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.root, text="Password").grid(row=1, column=0, padx=10, pady=10)
        self.entry_password = ttk.Entry(self.root, show="*")
        self.entry_password.grid(row=1, column=1, padx=10, pady=10)

        ttk.Button(self.root, text="Login", command=self.login).grid(row=2, column=0, columnspan=2, pady=10)

    def login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()
        if username == "ss" and password == "4176":
            self.root.destroy()
            root = tk.Tk()
            app = POSApp(root)
            root.mainloop()
        else:
            messagebox.showerror("Error", "Invalid credentials")

def add_product(file_name, product_id, name, price, quantity):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    ws.append([product_id, name, price, quantity])
    wb.save(file_name)

def display_products(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def search_product(file_name, search_term):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if search_term.lower() in str(row[0]).lower() or search_term.lower() in str(row[1]).lower():
            results.append(row)
    return results

def update_product(file_name, product_id, new_name=None, new_price=None, new_quantity=None):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            if new_name:
                row[1].value = new_name
            if new_price:
                row[2].value = new_price
            if new_quantity:
                row[3].value = new_quantity
            wb.save(file_name)
            return True
    return False

def delete_product(file_name, product_id):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            ws.delete_rows(row[0].row, 1)
            wb.save(file_name)
            return True
    return False

def place_order_product(file_name, product_id, order_quantity):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[3].value += order_quantity
            wb.save(file_name)
            return True
    return False

def purchase_product(file_name, product_id, purchase_quantity):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            if row[3].value >= purchase_quantity:
                row[3].value -= purchase_quantity
                wb.save(file_name)
                return True, row[3].value
            else:
                return False, row[3].value
    return False, None

class POSApp:
    def __init__(self, root):
        self.root = root
        self.root.title("POS System")
        self.root.geometry("800x600")
        self.root.configure(bg="#2C3E50")
        self.file_name = "products.xlsx"
        self.sales_file_name = "sales.xlsx"
        self.purchase_items = []
        self.cart = []
        self.create_widgets()

    def create_widgets(self):
        self.custom_font = font.Font(family="Helvetica", size=12, weight="bold")
        style = ttk.Style()
        style.configure("TNotebook", background="#34495E", foreground="#ECF0F1", font=self.custom_font)
        style.configure("TFrame", background="#34495E")
        style.configure("TLabel", background="#34495E", foreground="#ECF0F1", font=self.custom_font)
        style.configure("TEntry", font=self.custom_font)
        style.configure("Treeview.Heading", background="#34495E", foreground="#34495E",font=self.custom_font)  # Dark background with black text
        style.configure("Treeview", background="#ECF0F1", foreground="#34495E",font=self.custom_font)  # Set Treeview background and text colors
        style.configure("TButton", background="#2980B9", foreground="#34495E",font=self.custom_font, padding=5)  # Button with black text
        style.map("TButton", background=[("active", "#3498DB")])  # Change button color on hover

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=20, expand=True, fill=tk.BOTH)

        self.tab_welcome = ttk.Frame(self.notebook)
        self.tab_add = ttk.Frame(self.notebook)
        self.tab_view = ttk.Frame(self.notebook)
        self.tab_purchase = ttk.Frame(self.notebook)
        self.tab_admin = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_welcome, text="Welcome")
        self.notebook.add(self.tab_add, text="Add Product")
        self.notebook.add(self.tab_view, text="View Products")
        self.notebook.add(self.tab_purchase, text="Purchase Product")
        self.notebook.add(self.tab_admin, text="Admin")

        self.create_welcome_tab()
        self.create_add_product_tab()
        self.create_view_products_tab()
        self.create_purchase_product_tab()
        self.create_admin_tab()

    def create_welcome_tab(self):
        welcome_label = ttk.Label(self.tab_welcome, text="Welcome to SS INTERPRISORS POS System", font=("Helvetica", 24))
        welcome_label.pack(pady=20)
        footer_label = ttk.Label(self.tab_welcome, text="Contact: +92 333 5130796\nEmail: chsajjadshahid@outlook.com\nAddress: I-10/MARKAZ ,ISLAMABAD", font=("Helvetica", 12), background="#34495E", foreground="#ECF0F1")
        footer_label.pack(side="bottom", pady=20)

    def create_add_product_tab(self):
        ttk.Label(self.tab_add, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_product_id = ttk.Entry(self.tab_add)
        self.entry_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Name").grid(row=1, column=0, padx=10, pady=10)
        self.entry_product_name = ttk.Entry(self.tab_add)
        self.entry_product_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Price").grid(row=2, column=0, padx=10, pady=10)
        self.entry_product_price = ttk.Entry(self.tab_add)
        self.entry_product_price.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.tab_add, text="Quantity").grid(row=3, column=0, padx=10, pady=10)
        self.entry_product_quantity = ttk.Entry(self.tab_add)
        self.entry_product_quantity.grid(row=3, column=1, padx=10, pady=10)

        ttk.Button(self.tab_add, text="Add Product", command=self.add_product).grid(row=4, column=0, columnspan=2, pady=10)

    def add_product(self):
        product_id = int(self.entry_product_id.get())
        name = self.entry_product_name.get()
        price = float(self.entry_product_price.get())
        quantity = int(self.entry_product_quantity.get())

        add_product(self.file_name, product_id, name, price, quantity)
        messagebox.showinfo("Success", "Product added successfully")
        self.clear_entries()


    def create_view_products_tab(self):
        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_view)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Treeview for displaying products
        self.tree_view = ttk.Treeview(tree_frame, columns=("ID", "Name", "Price", "Quantity"), show="headings")
        self.tree_view.heading("ID", text="ID")
        self.tree_view.heading("Name", text="Name")
        self.tree_view.heading("Price", text="Price")
        self.tree_view.heading("Quantity", text="Quantity")
        self.tree_view.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_view.yview)
        self.tree_view.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_view.xview)
        self.tree_view.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Configure the grid layout for tree_frame to allow resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Refresh button
        ttk.Button(self.tab_view, text="Refresh", command=self.refresh_view_products).pack(pady=10)



    def refresh_view_products(self):
        for row in self.tree_view.get_children():
            self.tree_view.delete(row)
        for row in display_products(self.file_name):
            self.tree_view.insert("", "end", values=row)

    def create_purchase_product_tab(self):
        ttk.Label(self.tab_purchase, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_purchase_product_id = ttk.Entry(self.tab_purchase)
        self.entry_purchase_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_purchase, text="Quantity").grid(row=1, column=0, padx=10, pady=10)
        self.entry_purchase_quantity = ttk.Entry(self.tab_purchase)
        self.entry_purchase_quantity.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.tab_purchase, text="Discount (%)").grid(row=2, column=0, padx=10, pady=10)
        self.entry_discount = ttk.Entry(self.tab_purchase)
        self.entry_discount.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.tab_purchase, text="GST (%)").grid(row=3, column=0, padx=10, pady=10)
        self.entry_gst = ttk.Entry(self.tab_purchase)
        self.entry_gst.grid(row=3, column=1, padx=10, pady=10)

        ttk.Button(self.tab_purchase, text="Purchase", command=self.purchase_product).grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(self.tab_purchase, text="Generate Bill", command=self.create_bill_window).grid(row=5, column=0, columnspan=2, pady=10)

    def purchase_product(self):
        product_id = int(self.entry_purchase_product_id.get())
        quantity = int(self.entry_purchase_quantity.get())
        success, remaining = purchase_product(self.file_name, product_id, quantity)
        if success:
            discount = float(self.entry_discount.get()) if self.entry_discount.get() else 0
            gst = float(self.entry_gst.get()) if self.entry_gst.get() else 0
            messagebox.showinfo("Success", f"Purchased {quantity} items. Remaining stock: {remaining}")
            self.purchase_items.append((product_id, self.get_product_name(product_id), self.get_product_price(product_id), quantity, discount, gst))
        else:
            if remaining is not None:
                messagebox.showerror("Error", f"Not enough stock. Remaining stock: {remaining}")
            else:
                messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def get_product_name(self, product_id):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                return row[1]
        return ""

    def get_product_price(self, product_id):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                return row[2]
        return 0
#################################new purchase###############
    def create_admin_tab(self):
        self.admin_notebook = ttk.Notebook(self.tab_admin)
        self.admin_notebook.pack(pady=20, expand=True, fill=tk.BOTH)

        self.tab_search = ttk.Frame(self.admin_notebook)
        self.tab_update = ttk.Frame(self.admin_notebook)
        self.tab_delete = ttk.Frame(self.admin_notebook)
        self.tab_order = ttk.Frame(self.admin_notebook)
        self.tab_sales_report = ttk.Frame(self.admin_notebook)  # New Sales Report tab

        self.admin_notebook.add(self.tab_search, text="Search Product")
        self.admin_notebook.add(self.tab_update, text="Update Product")
        self.admin_notebook.add(self.tab_delete, text="Delete Product")
        self.admin_notebook.add(self.tab_order, text="Place Order")
        self.admin_notebook.add(self.tab_sales_report, text="Sales Report")  # Adding Sales Report tab

        self.create_search_product_tab()
        self.create_update_product_tab()
        self.create_delete_product_tab()
        self.create_order_product_tab()
        self.create_sales_report_tab()  # Call to create the sales report tab

    def create_sales_report_tab(self):
        # Month selection
        ttk.Label(self.tab_sales_report, text="Select Month", font=("Arial", 10, "bold"), foreground="white",
                  background="#2d3e50").grid(row=0, column=0, padx=10, pady=10, sticky="W")

        self.month_var = tk.StringVar()
        self.month_dropdown = ttk.Combobox(self.tab_sales_report, textvariable=self.month_var, width=15)
        self.month_dropdown['values'] = [month_name[i] for i in range(1, 13)]
        self.month_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky="W")

        # Year selection
        ttk.Label(self.tab_sales_report, text="Select Year", font=("Arial", 10, "bold"), foreground="white",
                  background="#2d3e50").grid(row=0, column=2, padx=10, pady=10, sticky="W")

        self.year_var = tk.StringVar()
        self.year_dropdown = ttk.Combobox(self.tab_sales_report, textvariable=self.year_var, width=15)
        self.year_dropdown['values'] = [str(year) for year in range(2000, datetime.now().year + 1)]
        self.year_dropdown.grid(row=0, column=3, padx=10, pady=10, sticky="W")

        # Filter Button
        filter_button = ttk.Button(self.tab_sales_report, text="Filter", command=self.filter_sales_report)
        filter_button.grid(row=0, column=4, padx=10, pady=10, sticky="W")

        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_sales_report)
        tree_frame.grid(row=1, column=0, columnspan=5, padx=10, pady=20, sticky="NSEW")

        # Treeview for displaying sales
        self.tree_sales = ttk.Treeview(tree_frame, columns=(
            "Date", "Product ID", "Product Name", "Quantity", "Price", "Total", "Discount", "GST", "Final Total"),
                                       show="headings")
        for col in self.tree_sales["columns"]:
            self.tree_sales.heading(col, text=col)

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_sales.yview)
        self.tree_sales.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_sales.xview)
        self.tree_sales.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Position the Treeview in the frame
        self.tree_sales.grid(row=0, column=0, sticky="nsew")

        # Configure tree_frame to be resizable
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure grid expansion for resizing
        self.tab_sales_report.columnconfigure(4, weight=1)
        self.tab_sales_report.rowconfigure(1, weight=1)

        # Load all sales initially
        self.load_sales_data()

    def load_sales_data(self):
        """Loads all sales data into the tree view"""
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)

        wb = openpyxl.load_workbook(self.sales_file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree_sales.insert("", "end", values=row)

    def filter_sales_report(self):
        """Filters sales based on selected month and year"""
        selected_month = self.month_var.get()
        selected_year = self.year_var.get()

        if not selected_month or not selected_year:
            messagebox.showerror("Error", "Please select both month and year.")
            return

        month_num = list(month_name).index(selected_month)  # Convert month name to month number
        filtered_sales = []

        wb = openpyxl.load_workbook(self.sales_file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            sale_date = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")

            if sale_date.month == month_num and sale_date.year == int(selected_year):
                filtered_sales.append(row)

        # Clear the tree view and insert filtered data
        for row in self.tree_sales.get_children():
            self.tree_sales.delete(row)
        for row in filtered_sales:
            self.tree_sales.insert("", "end", values=row)
    def create_search_product_tab(self):
        ttk.Label(self.tab_search, text="Search Term").grid(row=0, column=0, padx=10, pady=10)
        self.entry_search_term = ttk.Entry(self.tab_search)
        self.entry_search_term.grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(self.tab_search, text="Search", command=self.search_product).grid(row=1, column=0, columnspan=2,
                                                                                     pady=10)

        # Frame for Treeview and scrollbars
        tree_frame = ttk.Frame(self.tab_search)
        tree_frame.grid(row=2, column=0, columnspan=2, pady=10, padx=10, sticky="nsew")

        # Treeview for displaying search results
        self.tree_search = ttk.Treeview(tree_frame, columns=("ID", "Name", "Price", "Quantity"), show="headings")
        self.tree_search.heading("ID", text="ID")
        self.tree_search.heading("Name", text="Name")
        self.tree_search.heading("Price", text="Price")
        self.tree_search.heading("Quantity", text="Quantity")
        self.tree_search.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        tree_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_search.yview)
        self.tree_search.configure(yscrollcommand=tree_vsb.set)
        tree_vsb.grid(row=0, column=1, sticky="ns")

        # Horizontal scrollbar
        tree_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_search.xview)
        self.tree_search.configure(xscrollcommand=tree_hsb.set)
        tree_hsb.grid(row=1, column=0, sticky="ew")

        # Configure the grid layout for tree_frame to allow resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure the grid expansion for the main tab to allow resizing
        self.tab_search.columnconfigure(1, weight=1)
        self.tab_search.rowconfigure(2, weight=1)

    def search_product(self):
        search_term = self.entry_search_term.get()
        for row in self.tree_search.get_children():
            self.tree_search.delete(row)
        for row in search_product(self.file_name, search_term):
            self.tree_search.insert("", "end", values=row)

    def create_update_product_tab(self):
        ttk.Label(self.tab_update, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_update_product_id = ttk.Entry(self.tab_update)
        self.entry_update_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_update, text="New Name").grid(row=1, column=0, padx=10, pady=10)
        self.entry_update_name = ttk.Entry(self.tab_update)
        self.entry_update_name.grid(row=1, column=1, padx=10, pady=10)

        ttk.Label(self.tab_update, text="New Price").grid(row=2, column=0, padx=10, pady=10)
        self.entry_update_price = ttk.Entry(self.tab_update)
        self.entry_update_price.grid(row=2, column=1, padx=10, pady=10)

        ttk.Label(self.tab_update, text="New Quantity").grid(row=3, column=0, padx=10, pady=10)
        self.entry_update_quantity = ttk.Entry(self.tab_update)
        self.entry_update_quantity.grid(row=3, column=1, padx=10, pady=10)

        ttk.Button(self.tab_update, text="Update", command=self.update_product).grid(row=4, column=0, columnspan=2, pady=10)

    def update_product(self):
        product_id = int(self.entry_update_product_id.get())
        new_name = self.entry_update_name.get() or None
        new_price = self.entry_update_price.get()
        new_price = float(new_price) if new_price else None
        new_quantity = self.entry_update_quantity.get()
        new_quantity = int(new_quantity) if new_quantity else None

        if update_product(self.file_name, product_id, new_name, new_price, new_quantity):
            messagebox.showinfo("Success", "Product updated successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def create_delete_product_tab(self):
        ttk.Label(self.tab_delete, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_delete_product_id = ttk.Entry(self.tab_delete)
        self.entry_delete_product_id.grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(self.tab_delete, text="Delete", command=self.delete_product).grid(row=1, column=0, columnspan=2, pady=10)

    def delete_product(self):
        product_id = int(self.entry_delete_product_id.get())
        if delete_product(self.file_name, product_id):
            messagebox.showinfo("Success", "Product deleted successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def create_order_product_tab(self):
        ttk.Label(self.tab_order, text="Product ID").grid(row=0, column=0, padx=10, pady=10)
        self.entry_order_product_id = ttk.Entry(self.tab_order)
        self.entry_order_product_id.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(self.tab_order, text="Order Quantity").grid(row=1, column=0, padx=10, pady=10)
        self.entry_order_quantity = ttk.Entry(self.tab_order)
        self.entry_order_quantity.grid(row=1, column=1, padx=10, pady=10)

        ttk.Button(self.tab_order, text="Place Order", command=self.place_order_product).grid(row=2, column=0, columnspan=2, pady=10)

    def generate_bill(self):
        self.bill_text.delete(1.0, tk.END)
        bill_data = f"""
        SS INTERPRISORS
        Address: I-10/MARKAZ ,ISLAMABAD
        Contact: +92 333 5130796

        Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        {'-' * 40}\n"""

        total = 0
        final_total = 0
        for item in self.purchase_items:
            product_id, name, price, quantity, discount, gst = item
            line_total = price * quantity
            discount_amount = line_total * (discount / 100)
            line_total_after_discount = line_total - discount_amount
            gst_amount = line_total_after_discount * (gst / 100)
            final_line_total = line_total_after_discount + gst_amount

            bill_data += f"Name:{name}\t\tPrice:{price}\t\tQty:{quantity}\t\tTotal:{line_total:.2f}\t\tDISC:{discount:.2f}%\t\tGST:{gst:.2f}%\t\tFinaleTotal:{final_line_total:.2f}\n"
            total += line_total
            final_total += final_line_total

        bill_data += f"\n{'-' * 40}\nTotal:\t\t\t\t\t\t{total:.2f}\nFinal Total:\t\t\t\t\t\t{final_total:.2f}\n"

        self.bill_text.insert(tk.END, bill_data)
        self.record_sales()
    def record_sales(self):
        for item in self.purchase_items:
            product_id, name, price, quantity, discount, gst = item
            date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            line_total = price * quantity
            discount_amount = line_total * (discount / 100)
            line_total_after_discount = line_total - discount_amount
            gst_amount = line_total_after_discount * (gst / 100)
            final_line_total = line_total_after_discount + gst_amount
            sale_data = [date, product_id, name, quantity, price, line_total, discount, gst, final_line_total]
            record_sale(sale_data, self.sales_file_name)
        self.purchase_items = []

    def place_order_product(self):
        product_id = int(self.entry_order_product_id.get())
        order_quantity = int(self.entry_order_quantity.get())
        if place_order_product(self.file_name, product_id, order_quantity):
            messagebox.showinfo("Success", "Order placed successfully")
        else:
            messagebox.showerror("Error", "Product not found")
        self.clear_entries()

    def create_bill_window(self):
        self.bill_window = tk.Toplevel(self.root)
        self.bill_window.title("Bill Receipt")

        self.bill_text = tk.Text(self.bill_window, height=20, width=80)
        self.bill_text.pack(padx=20, pady=20)

        ttk.Button(self.bill_window, text="Print", command=self.print_bill).pack(pady=10)
        ttk.Button(self.bill_window, text="Generate Bill", command=self.generate_bill).pack(pady=10)

    def print_bill(self):
        bill_text = self.bill_text.get(1.0, tk.END)
        print(bill_text)
        # Optionally, you could implement printing to a physical printer using external libraries or OS-specific commands.

    def clear_entries(self):
        self.entry_product_id.delete(0, tk.END)
        self.entry_product_name.delete(0, tk.END)
        self.entry_product_price.delete(0, tk.END)
        self.entry_product_quantity.delete(0, tk.END)
        self.entry_purchase_product_id.delete(0, tk.END)
        self.entry_purchase_quantity.delete(0, tk.END)
        self.entry_discount.delete(0, tk.END)
        self.entry_gst.delete(0, tk.END)
        self.entry_search_term.delete(0, tk.END)
        self.entry_update_product_id.delete(0, tk.END)
        self.entry_update_name.delete(0, tk.END)
        self.entry_update_price.delete(0, tk.END)
        self.entry_update_quantity.delete(0, tk.END)
        self.entry_delete_product_id.delete(0, tk.END)
        self.entry_order_product_id.delete(0, tk.END)
        self.entry_order_quantity.delete(0, tk.END)

if __name__ == "__main__":
    create_product_excel()
    create_sales_excel()

    root = tk.Tk()
    app = LoginWindow(root)
    root.mainloop()

